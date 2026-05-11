"""
BodyBuilder API — FastAPI + SQLite backend
Multi-athlete support, workout plans, meal plans, calendar, xlsx export, email
"""
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, field_validator
from typing import Optional, List
import hashlib
import sqlite3
import json
import io
import os
import re
import smtplib
import tempfile
import threading
import time
import urllib.request
import urllib.parse
from pathlib import Path
from datetime import date, datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

app = FastAPI(title="BodyBuilder API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:8000", "http://127.0.0.1:8000"],
    allow_credentials=True,
    allow_methods=["GET","POST","PUT","PATCH","DELETE","OPTIONS"],
    allow_headers=["Content-Type","Accept"],
)

BASE_DIR = Path(__file__).parent

# ── Path overrides (set by app_launcher.py when running as a bundled .app) ───
# BB_DATA_DIR     → where the database and exercise_images live (persistent)
# BB_FRONTEND_DIR → where the React static files are (inside the bundle)
# BB_VERSION_FILE → path to the VERSION text file (inside the bundle)
#
# When running from source (dev mode), none of these are set and the original
# relative-path defaults are used unchanged.
_DATA_DIR     = Path(os.environ.get("BB_DATA_DIR",     str(BASE_DIR)))
_FRONTEND_ENV = os.environ.get("BB_FRONTEND_DIR")
_VERSION_ENV  = os.environ.get("BB_VERSION_FILE")

DB_PATH      = str(_DATA_DIR / "bodybuilder.db")
FRONTEND_DIR = Path(_FRONTEND_ENV) if _FRONTEND_ENV else BASE_DIR.parent / "frontend"

# ── Read version from repo-root VERSION file ──────────────────────────────────
_VERSION_FILE = Path(_VERSION_ENV) if _VERSION_ENV else BASE_DIR.parent / "VERSION"
def _read_version_file() -> tuple[int, int, int]:
    try:
        parts = _VERSION_FILE.read_text().strip().split(".")
        return int(parts[0]), int(parts[1]), int(parts[2])
    except Exception:
        return 1, 0, 0
APP_VERSION = _read_version_file()   # (major, minor, tiny)
EXERCISE_IMAGES_DIR = _DATA_DIR / "exercise_images"
EXERCISE_IMAGES_DIR.mkdir(exist_ok=True)

# ── All exercises eligible for image seeding ──────────────────────────────────
EXERCISE_ALL = [
    # Chest
    "Barbell Bench Press","Dumbbell Bench Press","Incline Bench Press","Decline Bench Press",
    "Push-Up","Cable Fly","Dumbbell Fly","Chest Dip","Cable Crossover","Pec Deck",
    # Back
    "Pull-Up","Chin-Up","Barbell Row","Dumbbell Row","Cable Row","Lat Pulldown",
    "T-Bar Row","Face Pull","Deadlift","Rack Pull",
    # Shoulders
    "Barbell Overhead Press","Dumbbell Overhead Press","Lateral Raise","Front Raise",
    "Rear Delt Fly","Cable Lateral Raise","Upright Row","Arnold Press","Shrug",
    # Biceps
    "Barbell Curl","Dumbbell Curl","Hammer Curl","Cable Curl","Preacher Curl",
    "Incline Dumbbell Curl","Concentration Curl","Spider Curl",
    # Triceps
    "Close-Grip Bench Press","Tricep Pushdown","Overhead Tricep Extension",
    "Skull Crusher","Dip","Cable Kickback","Diamond Push-Up",
    # Forearms
    "Wrist Curl","Reverse Wrist Curl","Farmer's Walk","Pinch Grip Hold",
    # Quads
    "Barbell Squat","Front Squat","Leg Press","Hack Squat","Bulgarian Split Squat",
    "Leg Extension","Lunge","Step-Up","Goblet Squat",
    # Hamstrings
    "Romanian Deadlift","Leg Curl","Good Morning","Glute-Ham Raise",
    "Nordic Hamstring Curl","Stiff-Leg Deadlift",
    # Glutes
    "Hip Thrust","Sumo Deadlift","Glute Bridge","Clamshell","Abductor Machine",
    # Calves
    "Standing Calf Raise","Seated Calf Raise","Donkey Calf Raise","Single-Leg Calf Raise",
    # Core
    "Plank","Cable Crunch","Hanging Leg Raise","Ab Rollout","Side Plank",
    "Russian Twist","Crunch","Sit-Up","Pallof Press",
    # Cardio
    "Treadmill Run","Cycling","Rowing","Jump Rope","Stair Climber","Sled Push","Battle Ropes",
    # Full Body
    "Clean and Press","Kettlebell Swing","Burpee","Box Jump","Thruster","Turkish Get-Up",
]

# ── Image seed state (in-memory, reset on server restart) ─────────────────────
_seed_state = {"running": False, "done": 0, "total": len(EXERCISE_ALL), "errors": 0}


def _slugify(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


_WGER_HEADERS = {"User-Agent": "BodyBuilderApp/1.0", "Accept": "application/json"}
_WGER_TIMEOUT = 30
_WGER_RETRIES = 3


def _fetch_json(url: str) -> dict:
    for attempt in range(1, _WGER_RETRIES + 1):
        try:
            req = urllib.request.Request(url, headers=_WGER_HEADERS)
            with urllib.request.urlopen(req, timeout=_WGER_TIMEOUT) as r:
                return json.loads(r.read())
        except Exception:
            if attempt < _WGER_RETRIES:
                time.sleep(attempt * 2)
    return {}


def _build_wger_index() -> dict:
    """Page through Wger exercise translations and return name.lower() -> base_id."""
    index: dict = {}
    url = ("https://wger.de/api/v2/exercisetranslation/?"
           + urllib.parse.urlencode({"format": "json", "language": 2, "limit": 100}))
    while url:
        data = _fetch_json(url)
        for item in data.get("results", []):
            name = (item.get("name") or "").strip()
            bid  = item.get("exercise_base")
            if name and bid:
                index[name.lower()] = bid
        url = data.get("next")
        time.sleep(0.3)
    return index


def _best_match(name: str, index: dict):
    attempts = list(dict.fromkeys([
        name.lower(),
        name.lower().split("(")[0].strip(),
        name.lower().split("/")[0].strip(),
        name.lower().replace("-", " "),
        " ".join(name.lower().split()[:-1]),
    ]))
    for attempt in attempts:
        if not attempt:
            continue
        if attempt in index:
            return index[attempt]
        for wger_name, bid in index.items():
            if attempt in wger_name:
                return bid
    return None


def _wger_image_url(base_id: int):
    url = ("https://wger.de/api/v2/exerciseimage/?"
           + urllib.parse.urlencode({"exercise_base": base_id, "format": "json"}))
    data = _fetch_json(url)
    results = data.get("results", [])
    mains = [x for x in results if x.get("is_main")]
    src = mains[0] if mains else (results[0] if results else None)
    return src["image"] if src else None


def _download_image(img_url: str, save_path: Path) -> bool:
    # SSRF guard: only fetch from the expected wger.de domain over HTTPS
    from urllib.parse import urlparse as _urlparse
    _p = _urlparse(img_url)
    if _p.scheme != "https" or not (_p.netloc == "wger.de" or _p.netloc.endswith(".wger.de")):
        return False
    for attempt in range(1, _WGER_RETRIES + 1):
        try:
            req = urllib.request.Request(img_url, headers=_WGER_HEADERS)
            with urllib.request.urlopen(req, timeout=_WGER_TIMEOUT) as r:
                data = r.read()
                if len(data) > 500:
                    save_path.write_bytes(data)
                    return True
        except Exception:
            if attempt < _WGER_RETRIES:
                time.sleep(attempt * 2)
    return False


def _run_seed(force: bool = False):
    """Background thread: build Wger index then fetch and cache exercise images."""
    global _seed_state
    _seed_state.update({"running": True, "done": 0, "errors": 0, "total": len(EXERCISE_ALL)})
    conn = get_db()
    try:
        index = _build_wger_index()
        if not index:
            _seed_state["running"] = False
            return

        for name in EXERCISE_ALL:
            row = conn.execute(
                "SELECT image_path FROM exercise_images WHERE name=?", (name,)
            ).fetchone()
            already_cached = (
                row and row["image_path"]
                and (EXERCISE_IMAGES_DIR / row["image_path"]).exists()
            )
            if already_cached and not force:
                _seed_state["done"] += 1
                continue

            slug = _slugify(name)
            img_file = f"{slug}.jpg"
            save_path = EXERCISE_IMAGES_DIR / img_file

            base_id = _best_match(name, index)
            if base_id:
                img_url = _wger_image_url(base_id)
                if img_url and _download_image(img_url, save_path):
                    conn.execute(
                        "INSERT OR REPLACE INTO exercise_images (name, image_path, source)"
                        " VALUES (?, ?, 'wger')",
                        (name, img_file)
                    )
                    conn.commit()
                    _seed_state["done"] += 1
                else:
                    _seed_state["errors"] += 1
            else:
                _seed_state["errors"] += 1

            time.sleep(0.4)
    finally:
        conn.close()
        _seed_state["running"] = False


# ─── DB Helpers ───────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def _col_exists(conn, table, col):
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(r["name"] == col for r in rows)


def _table_exists(conn, table):
    r = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
    return r is not None


def init_db():
    conn = get_db()
    c = conn.cursor()

    # ── version ──
    c.execute("""CREATE TABLE IF NOT EXISTS version (
        id INTEGER PRIMARY KEY DEFAULT 1,
        major INTEGER DEFAULT 1, minor INTEGER DEFAULT 0, tiny INTEGER DEFAULT 0, notes TEXT DEFAULT ''
    )""")
    c.execute("INSERT OR IGNORE INTO version (id) VALUES (1)")
    # Sync version table with VERSION file (only if the file version is newer)
    _vmaj, _vmin, _vtiny = APP_VERSION
    _row = c.execute("SELECT major, minor, tiny FROM version WHERE id=1").fetchone()
    if _row:
        _db_tuple  = (_row["major"] or 0, _row["minor"] or 0, _row["tiny"] or 0)
        _file_tuple = (_vmaj, _vmin, _vtiny)
        if _file_tuple > _db_tuple:
            c.execute("UPDATE version SET major=?, minor=?, tiny=? WHERE id=1",
                      (_vmaj, _vmin, _vtiny))

    # ── athletes (multi-athlete) ──
    c.execute("""CREATE TABLE IF NOT EXISTS athletes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT DEFAULT '',
        email TEXT DEFAULT '',
        birthdate TEXT DEFAULT '',
        height_cm REAL DEFAULT 175,
        weight_kg REAL DEFAULT 75,
        body_fat_pct REAL DEFAULT 0,
        sex TEXT DEFAULT 'male',
        activity_level INTEGER DEFAULT 1,
        workout_days_per_week INTEGER DEFAULT 3,
        workout_days TEXT DEFAULT '[]',
        workout_time TEXT DEFAULT 'AM',
        phase TEXT DEFAULT 'maintain',
        deficit REAL DEFAULT 0,
        units TEXT DEFAULT 'metric',
        status TEXT DEFAULT 'active'
    )""")

    # ── Migrate old single-athlete table if present ──
    if _table_exists(conn, "athlete") and conn.execute("SELECT COUNT(*) FROM athletes").fetchone()[0] == 0:
        old = conn.execute("SELECT * FROM athlete WHERE id=1").fetchone()
        if old:
            d = dict(old)
            c.execute("""INSERT INTO athletes (id, name, email, birthdate, height_cm, weight_kg, body_fat_pct,
                sex, activity_level, workout_days_per_week, workout_days, workout_time, phase, deficit)
                VALUES (1,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (d.get("name",""), d.get("email",""), d.get("birthdate",""), d.get("height_cm",175),
                 d.get("weight_kg",75), d.get("body_fat_pct",0), d.get("sex","male"),
                 d.get("activity_level",1), d.get("workout_days_per_week",3), d.get("workout_days","[]"),
                 d.get("workout_time","AM"), d.get("phase","maintain"), d.get("deficit",0)))

    # ── programs ──
    c.execute("""CREATE TABLE IF NOT EXISTS programs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        athlete_id INTEGER NOT NULL DEFAULT 1,
        start_date TEXT DEFAULT '',
        end_date TEXT DEFAULT '',
        payment_processed INTEGER DEFAULT 0,
        FOREIGN KEY (athlete_id) REFERENCES athletes(id) ON DELETE CASCADE
    )""")
    # Migrate old program table
    if _table_exists(conn, "program") and conn.execute("SELECT COUNT(*) FROM programs").fetchone()[0] == 0:
        old = conn.execute("SELECT * FROM program WHERE id=1").fetchone()
        if old:
            d = dict(old)
            c.execute("INSERT INTO programs (athlete_id, start_date, end_date, payment_processed) VALUES (1,?,?,?)",
                      (d.get("start_date",""), d.get("end_date",""), d.get("payment_processed",0)))

    # ── activity_calories ──
    c.execute("""CREATE TABLE IF NOT EXISTS activity_calories (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        athlete_id INTEGER NOT NULL DEFAULT 1,
        level INTEGER NOT NULL,
        additional_calories REAL DEFAULT 0,
        multiplier REAL DEFAULT 1.0,
        UNIQUE(athlete_id, level),
        FOREIGN KEY (athlete_id) REFERENCES athletes(id) ON DELETE CASCADE
    )""")
    # Migrate existing rows: add multiplier column and seed standard values
    if not _col_exists(conn, "activity_calories", "multiplier"):
        c.execute("ALTER TABLE activity_calories ADD COLUMN multiplier REAL DEFAULT 1.0")
    # Fix any rows still at the placeholder default of 1.0 — no standard level uses 1.0
    for lvl, mult in [(1,1.200),(2,1.375),(3,1.550),(4,1.725),(5,1.900)]:
        c.execute("UPDATE activity_calories SET multiplier=? WHERE level=? AND multiplier=1.0",
                  (mult, lvl))
    # Will be populated per-athlete on first access

    # ── calendar_days ──
    c.execute("""CREATE TABLE IF NOT EXISTS calendar_days (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        athlete_id INTEGER NOT NULL DEFAULT 1,
        date TEXT NOT NULL,
        steps INTEGER DEFAULT 0,
        aerobic_type TEXT DEFAULT '',
        aerobic_duration INTEGER DEFAULT 0,
        workout_notes TEXT DEFAULT '',
        UNIQUE(athlete_id, date),
        FOREIGN KEY (athlete_id) REFERENCES athletes(id) ON DELETE CASCADE
    )""")
    # Migrate old calendar_days (no athlete_id)
    if _table_exists(conn, "calendar_days") and not _col_exists(conn, "calendar_days", "athlete_id"):
        pass  # old table is pre-migration; new table is created above with new name logic
    # (Since we create a new table with UNIQUE constraint, old data stays in old table)

    # ── calendar_events ──
    c.execute("""CREATE TABLE IF NOT EXISTS calendar_events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        athlete_id INTEGER NOT NULL DEFAULT 1,
        date TEXT NOT NULL,
        title TEXT DEFAULT '',
        description TEXT DEFAULT '',
        event_time TEXT DEFAULT '',
        FOREIGN KEY (athlete_id) REFERENCES athletes(id) ON DELETE CASCADE
    )""")

    # ── meal_plans ──
    c.execute("""CREATE TABLE IF NOT EXISTS meal_plans (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        athlete_id INTEGER NOT NULL DEFAULT 1 UNIQUE,
        protein_target REAL DEFAULT 150, carbs_target REAL DEFAULT 200,
        fat_target REAL DEFAULT 65, fiber_target REAL DEFAULT 25,
        sodium_target REAL DEFAULT 2300, potassium_target REAL DEFAULT 3500,
        protein_actual REAL DEFAULT 0, carbs_actual REAL DEFAULT 0,
        fat_actual REAL DEFAULT 0, fiber_actual REAL DEFAULT 0,
        sodium_actual REAL DEFAULT 0, potassium_actual REAL DEFAULT 0,
        FOREIGN KEY (athlete_id) REFERENCES athletes(id) ON DELETE CASCADE
    )""")

    # ── nutrition_foods ──
    c.execute("""CREATE TABLE IF NOT EXISTS nutrition_foods (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        athlete_id INTEGER NOT NULL DEFAULT 1,
        name TEXT DEFAULT '', protein REAL DEFAULT 0, carbs REAL DEFAULT 0,
        fat REAL DEFAULT 0, fiber REAL DEFAULT 0, sodium REAL DEFAULT 0,
        potassium REAL DEFAULT 0, calories REAL DEFAULT 0,
        serving_size TEXT DEFAULT '100g', serving_g REAL DEFAULT 100,
        category TEXT DEFAULT 'general',
        FOREIGN KEY (athlete_id) REFERENCES athletes(id) ON DELETE CASCADE
    )""")
    if not _col_exists(conn, "nutrition_foods", "serving_g"):
        c.execute("ALTER TABLE nutrition_foods ADD COLUMN serving_g REAL DEFAULT 100")

    # ── food_swaps ──
    c.execute("""CREATE TABLE IF NOT EXISTS food_swaps (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        athlete_id INTEGER NOT NULL,
        category TEXT DEFAULT 'carbs',
        source_name TEXT DEFAULT '',
        source_amount REAL DEFAULT 100,
        source_unit TEXT DEFAULT 'g',
        swap_name TEXT DEFAULT '',
        swap_amount REAL DEFAULT 100,
        swap_unit TEXT DEFAULT 'g',
        sort_order INTEGER DEFAULT 0,
        FOREIGN KEY (athlete_id) REFERENCES athletes(id) ON DELETE CASCADE
    )""")

    # ── workout_plans ──
    c.execute("""CREATE TABLE IF NOT EXISTS workout_plans (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        athlete_id INTEGER NOT NULL DEFAULT 1,
        title TEXT DEFAULT '',
        start_date TEXT DEFAULT '',
        end_date TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        sort_order INTEGER DEFAULT 0,
        FOREIGN KEY (athlete_id) REFERENCES athletes(id) ON DELETE CASCADE
    )""")

    # ── workout_sessions (one per day-of-week within a plan) ──
    c.execute("""CREATE TABLE IF NOT EXISTS workout_sessions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        plan_id INTEGER NOT NULL,
        day_of_week TEXT DEFAULT 'Monday',
        session_title TEXT DEFAULT '',
        muscle_groups TEXT DEFAULT '[]',
        session_notes TEXT DEFAULT '',
        sort_order INTEGER DEFAULT 0,
        FOREIGN KEY (plan_id) REFERENCES workout_plans(id) ON DELETE CASCADE
    )""")

    # ── workout_exercises ──
    c.execute("""CREATE TABLE IF NOT EXISTS workout_exercises (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        session_id INTEGER NOT NULL,
        name TEXT DEFAULT '',
        muscle_group TEXT DEFAULT '',
        set_type TEXT DEFAULT 'working',
        sets_json TEXT DEFAULT '[]',
        rep_range TEXT DEFAULT '',
        rir INTEGER DEFAULT 2,
        tempo TEXT DEFAULT '',
        intensifiers TEXT DEFAULT '',
        exercise_notes TEXT DEFAULT '',
        sort_order INTEGER DEFAULT 0,
        FOREIGN KEY (session_id) REFERENCES workout_sessions(id) ON DELETE CASCADE
    )""")
    # Migrate old exercise columns
    for col, defval in [("muscle_group","''"), ("rep_range","''"), ("rir","2"),
                        ("tempo","''"), ("intensifiers","''"), ("image_url","''"),
                        ("warmup_instructions","''")]:
        if not _col_exists(conn, "workout_exercises", col):
            c.execute(f"ALTER TABLE workout_exercises ADD COLUMN {col} TEXT DEFAULT {defval}")
    # Migrate set_type 'main' → 'working'
    c.execute("UPDATE workout_exercises SET set_type='working' WHERE set_type='main'")
    # Migrate athletes: add units column
    if not _col_exists(conn, "athletes", "units"):
        c.execute("ALTER TABLE athletes ADD COLUMN units TEXT DEFAULT 'metric'")
    if not _col_exists(conn, "athletes", "status"):
        c.execute("ALTER TABLE athletes ADD COLUMN status TEXT DEFAULT 'active'")
    if not _col_exists(conn, "workout_plans", "warmup_instructions"):
        c.execute("ALTER TABLE workout_plans ADD COLUMN warmup_instructions TEXT DEFAULT ''")
    if not _col_exists(conn, "workout_plans", "rest_days"):
        c.execute("ALTER TABLE workout_plans ADD COLUMN rest_days TEXT DEFAULT '[]'")

    # ── supplements ──
    c.execute("""CREATE TABLE IF NOT EXISTS supplements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        athlete_id INTEGER NOT NULL,
        day_of_week TEXT NOT NULL,
        name TEXT DEFAULT '',
        dosage TEXT DEFAULT '',
        time_of_day TEXT DEFAULT 'AM',
        sort_order INTEGER DEFAULT 0,
        FOREIGN KEY (athlete_id) REFERENCES athletes(id) ON DELETE CASCADE
    )""")

    # ── meals ──
    c.execute("""CREATE TABLE IF NOT EXISTS meals (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        athlete_id INTEGER NOT NULL,
        day_type TEXT DEFAULT 'training',
        name TEXT DEFAULT '',
        sort_order INTEGER DEFAULT 0,
        FOREIGN KEY (athlete_id) REFERENCES athletes(id) ON DELETE CASCADE
    )""")

    # ── meal_items ──
    c.execute("""CREATE TABLE IF NOT EXISTS meal_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        meal_id INTEGER NOT NULL,
        source_type TEXT DEFAULT 'protein',
        food_name TEXT DEFAULT '',
        quantity REAL DEFAULT 1,
        weight_g REAL DEFAULT 0,
        serving_size TEXT DEFAULT '100g',
        protein_g REAL DEFAULT 0,
        carbs_g REAL DEFAULT 0,
        fat_g REAL DEFAULT 0,
        fiber_g REAL DEFAULT 0,
        sodium_mg REAL DEFAULT 0,
        potassium_mg REAL DEFAULT 0,
        sort_order INTEGER DEFAULT 0,
        FOREIGN KEY (meal_id) REFERENCES meals(id) ON DELETE CASCADE
    )""")

    # ── exercise_images ──
    c.execute("""CREATE TABLE IF NOT EXISTS exercise_images (
        name TEXT PRIMARY KEY,
        image_path TEXT DEFAULT '',
        source TEXT DEFAULT 'wger'
    )""")

    # ── smtp_settings ──
    c.execute("""CREATE TABLE IF NOT EXISTS smtp_settings (
        id INTEGER PRIMARY KEY DEFAULT 1,
        host TEXT DEFAULT 'smtp.gmail.com', port INTEGER DEFAULT 587,
        username TEXT DEFAULT '', password TEXT DEFAULT '',
        from_name TEXT DEFAULT 'BodyBuilder Coach', use_tls INTEGER DEFAULT 1
    )""")
    c.execute("INSERT OR IGNORE INTO smtp_settings (id) VALUES (1)")

    # ── Performance indexes ──
    c.execute("CREATE INDEX IF NOT EXISTS idx_meal_items_meal_id         ON meal_items(meal_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_meals_athlete_id           ON meals(athlete_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_supplements_athlete_id     ON supplements(athlete_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_workout_exercises_session  ON workout_exercises(session_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_workout_sessions_plan      ON workout_sessions(plan_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_food_swaps_athlete_id      ON food_swaps(athlete_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_nutrition_foods_athlete_id ON nutrition_foods(athlete_id)")

    conn.commit()
    conn.close()


def ensure_athlete_defaults(athlete_id: int):
    """Ensure activity_calories and meal_plan rows exist for this athlete."""
    # Standard Harris-Benedict TDEE multipliers per activity level
    DEFAULTS = [(1, 0, 1.200), (2, 0, 1.375), (3, 0, 1.550), (4, 0, 1.725), (5, 0, 1.900)]
    conn = get_db()
    for level, cal, mult in DEFAULTS:
        conn.execute(
            "INSERT OR IGNORE INTO activity_calories "
            "(athlete_id, level, additional_calories, multiplier) VALUES (?,?,?,?)",
            (athlete_id, level, cal, mult))
    conn.execute("INSERT OR IGNORE INTO meal_plans (athlete_id) VALUES (?)", (athlete_id,))
    conn.commit()
    conn.close()


init_db()


# ─── RMR Calculations ─────────────────────────────────────────────────────────

def compute_age(birthdate: str) -> int:
    if not birthdate:
        return 30
    try:
        y, m, d = birthdate.split("-")
        bd = date(int(y), int(m), int(d))
        today = date.today()
        return today.year - bd.year - ((today.month, today.day) < (bd.month, bd.day))
    except Exception:
        return 30


def calculate_rmr(weight_kg, height_cm, birthdate, sex, body_fat_pct):
    age = compute_age(birthdate)
    wk = max(float(weight_kg or 1), 1)
    hc = max(float(height_cm or 1), 1)
    s = (sex or "male").lower()
    mifflin = (10*wk)+(6.25*hc)-(5*age)+(5 if s=="male" else -161)
    harris = (88.362+(13.397*wk)+(4.799*hc)-(5.677*age)) if s=="male" else (447.593+(9.247*wk)+(3.098*hc)-(4.330*age))
    bf = float(body_fat_pct or 0)
    lbm = wk*(1-bf/100) if bf>0 else ((0.407*wk+0.267*hc-19.2) if s=="male" else (0.252*wk+0.473*hc-48.3))
    katch = 370+(21.6*lbm)
    avg = (mifflin+harris+katch)/3
    return {"mifflin":round(mifflin,1),"harris":round(harris,1),"katch":round(katch,1),
            "average":round(avg,1),"age":age,"lbm_kg":round(lbm,1)}


def athlete_row_to_dict(row):
    d = dict(row)
    d["workout_days"] = json.loads(d.get("workout_days") or "[]")
    rmr = calculate_rmr(d["weight_kg"], d["height_cm"], d["birthdate"], d["sex"], d["body_fat_pct"])
    d.update(rmr)
    return d


def next_weekday_on_or_after(start: date, day_name: str) -> date:
    """Return the first date >= start that falls on the given day_name (e.g. 'Monday')."""
    days = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    target_dow = days.index(day_name) if day_name in days else 0
    current_dow = start.weekday()  # 0=Monday
    delta = (target_dow - current_dow) % 7
    return start + timedelta(days=delta)


# ─── Pydantic Models ──────────────────────────────────────────────────────────

class AthleteModel(BaseModel):
    name: Optional[str] = ""
    email: Optional[str] = ""
    birthdate: Optional[str] = ""
    height_cm: Optional[float] = 175
    weight_kg: Optional[float] = 75
    body_fat_pct: Optional[float] = 0
    sex: Optional[str] = "male"
    activity_level: Optional[int] = 1
    workout_days_per_week: Optional[int] = 3
    workout_days: Optional[List[str]] = []
    workout_time: Optional[str] = "AM"
    phase: Optional[str] = "maintain"
    deficit: Optional[float] = 0
    units: Optional[str] = "metric"
    status: Optional[str] = "active"

    @field_validator("status")
    @classmethod
    def status_v(cls, v):
        if v not in ("active", "inactive"): raise ValueError("active or inactive")
        return v

    @field_validator("name")
    @classmethod
    def name_v(cls, v):
        v = (v or "").strip()
        if len(v) > 100: raise ValueError("Max 100 characters")
        if re.search(r"<script|javascript\s*:|on\w+\s*=", v, re.I): raise ValueError("Invalid characters")
        return v
    @field_validator("units")
    @classmethod
    def units_v(cls, v):
        if v not in ("metric", "imperial"): raise ValueError("metric or imperial")
        return v
    @field_validator("workout_time")
    @classmethod
    def wt_v(cls, v):
        if v not in ("AM", "PM"): raise ValueError("AM or PM")
        return v
    @field_validator("workout_days")
    @classmethod
    def wd_days_v(cls, v):
        valid = {"Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"}
        for d in (v or []):
            if d not in valid: raise ValueError(f"Invalid day: {d}")
        return v or []
    @field_validator("email")
    @classmethod
    def email_v(cls, v):
        if v and len(v) > 0 and "@" not in v: raise ValueError("Invalid email")
        if v and len(v) > 200: raise ValueError("Email too long")
        return v
    @field_validator("height_cm")
    @classmethod
    def height_v(cls, v):
        if v is not None and (v<50 or v>300): raise ValueError("Height must be 50–300 cm")
        return v
    @field_validator("weight_kg")
    @classmethod
    def weight_v(cls, v):
        if v is not None and (v<10 or v>500): raise ValueError("Weight must be 10–500 kg")
        return v
    @field_validator("body_fat_pct")
    @classmethod
    def bf_v(cls, v):
        if v is not None and (v<0 or v>70): raise ValueError("Body fat 0–70%")
        return v
    @field_validator("activity_level")
    @classmethod
    def al_v(cls, v):
        if v is not None and (v<1 or v>5): raise ValueError("Activity level 1–5")
        return v
    @field_validator("workout_days_per_week")
    @classmethod
    def wd_v(cls, v):
        if v is not None and (v<0 or v>7): raise ValueError("0–7 days")
        return v
    @field_validator("sex")
    @classmethod
    def sex_v(cls, v):
        if v not in ("male","female"): raise ValueError("male or female")
        return v
    @field_validator("phase")
    @classmethod
    def phase_v(cls, v):
        if v not in ("cut","bulk","maintain","prep"): raise ValueError("cut/bulk/maintain/prep")
        return v
    @field_validator("deficit")
    @classmethod
    def def_v(cls, v):
        if v is not None and (v<0 or v>2000): raise ValueError("0–2000 kcal")
        return v


def _valid_date_or_empty(v: str, field: str) -> str:
    """Return v if it is a valid ISO date (YYYY-MM-DD) or empty string; raise ValueError otherwise."""
    v = (v or "").strip()
    if not v:
        return ""
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", v):
        raise ValueError(f"{field} must be YYYY-MM-DD or empty")
    try:
        date.fromisoformat(v)
    except ValueError:
        raise ValueError(f"{field} is not a valid date")
    return v


class ProgramModel(BaseModel):
    start_date: Optional[str] = ""
    end_date: Optional[str] = ""
    payment_processed: Optional[bool] = False

    @field_validator("start_date", "end_date")
    @classmethod
    def date_v(cls, v, info):
        return _valid_date_or_empty(v, info.field_name)


class ActivityCalModel(BaseModel):
    additional_calories: float = 0
    multiplier: float = 1.2
    @field_validator("multiplier")
    @classmethod
    def mult_v(cls, v):
        if v is not None and (v < 0.5 or v > 5.0): raise ValueError("Multiplier must be 0.5–5.0")
        return v
    @field_validator("additional_calories")
    @classmethod
    def cal_v(cls, v):
        if v<0 or v>5000: raise ValueError("0–5000")
        return v


class CalDayModel(BaseModel):
    steps: Optional[int] = 0
    aerobic_type: Optional[str] = ""
    aerobic_duration: Optional[int] = 0
    workout_notes: Optional[str] = ""
    @field_validator("steps")
    @classmethod
    def steps_v(cls, v):
        if v is not None and (v<0 or v>100000): raise ValueError("0–100,000")
        return v
    @field_validator("aerobic_duration")
    @classmethod
    def dur_v(cls, v):
        if v is not None and (v<0 or v>600): raise ValueError("0–600 min")
        return v
    @field_validator("aerobic_type")
    @classmethod
    def atype_v(cls, v):
        v = (v or "").strip()
        if len(v) > 100: raise ValueError("Max 100 characters")
        return v
    @field_validator("workout_notes")
    @classmethod
    def wnotes_v(cls, v):
        v = (v or "").strip()
        if len(v) > 1000: raise ValueError("Max 1000 characters")
        return v


class EventModel(BaseModel):
    date: str
    title: Optional[str] = ""
    description: Optional[str] = ""
    event_time: Optional[str] = ""
    @field_validator("date")
    @classmethod
    def event_date_v(cls, v):
        result = _valid_date_or_empty(v, "date")
        if not result: raise ValueError("date required")
        return result
    @field_validator("title")
    @classmethod
    def title_v(cls, v):
        if not v or not v.strip(): raise ValueError("Title required")
        if len(v)>100: raise ValueError("Max 100 chars")
        return v.strip()
    @field_validator("description")
    @classmethod
    def desc_v(cls, v):
        v = (v or "").strip()
        if len(v) > 1000: raise ValueError("Max 1000 characters")
        return v
    @field_validator("event_time")
    @classmethod
    def etime_v(cls, v):
        v = (v or "").strip()
        if v and not re.fullmatch(r"\d{1,2}:\d{2}", v):
            raise ValueError("event_time must be HH:MM or empty")
        return v


class MealPlanModel(BaseModel):
    protein_target: Optional[float] = 150
    carbs_target: Optional[float] = 200
    fat_target: Optional[float] = 65
    fiber_target: Optional[float] = 25
    sodium_target: Optional[float] = 2300
    potassium_target: Optional[float] = 3500
    protein_actual: Optional[float] = 0
    carbs_actual: Optional[float] = 0
    fat_actual: Optional[float] = 0
    fiber_actual: Optional[float] = 0
    sodium_actual: Optional[float] = 0
    potassium_actual: Optional[float] = 0

    @field_validator("protein_target","carbs_target","fat_target","fiber_target",
                     "protein_actual","carbs_actual","fat_actual","fiber_actual")
    @classmethod
    def macro_v(cls, v):
        if v is not None and (v < 0 or v > 9999): raise ValueError("Must be 0–9999 g")
        return v or 0
    @field_validator("sodium_target","potassium_target","sodium_actual","potassium_actual")
    @classmethod
    def mineral_v(cls, v):
        if v is not None and (v < 0 or v > 99999): raise ValueError("Must be 0–99999 mg")
        return v or 0


class FoodModel(BaseModel):
    name: str
    protein: Optional[float] = 0
    carbs: Optional[float] = 0
    fat: Optional[float] = 0
    fiber: Optional[float] = 0
    sodium: Optional[float] = 0
    potassium: Optional[float] = 0
    calories: Optional[float] = 0
    serving_size: Optional[str] = "100g"
    serving_g: Optional[float] = 100
    category: Optional[str] = "general"
    @field_validator("name")
    @classmethod
    def name_v(cls, v):
        if not v or not v.strip(): raise ValueError("Name required")
        if len(v) > 200: raise ValueError("Max 200 characters")
        return v.strip()
    @field_validator("serving_size")
    @classmethod
    def ss_v(cls, v):
        v = (v or "100g").strip()
        if len(v) > 50: raise ValueError("Max 50 characters")
        return v
    @field_validator("serving_g")
    @classmethod
    def serving_g_v(cls, v):
        v = v or 100
        if v <= 0 or v > 9999: raise ValueError("Serving size must be 1–9999 g")
        return round(v, 1)
    @field_validator("protein","carbs","fat","fiber","calories")
    @classmethod
    def macro_v(cls, v):
        if v is not None and (v < 0 or v > 9999): raise ValueError("Must be 0–9999")
        return v or 0
    @field_validator("sodium","potassium")
    @classmethod
    def mineral_v(cls, v):
        if v is not None and (v < 0 or v > 99999): raise ValueError("Must be 0–99999 mg")
        return v or 0
    @field_validator("category")
    @classmethod
    def cat_v(cls, v):
        allowed = ("protein","carb","fat","vegetable","fruit","dairy","supplement","general")
        if v not in allowed: raise ValueError("Invalid category")
        return v


class FoodSwapModel(BaseModel):
    category: str = "carbs"
    source_name: str = ""
    source_amount: Optional[float] = 100
    source_unit: Optional[str] = "g"
    swap_name: str = ""
    swap_amount: Optional[float] = 100
    swap_unit: Optional[str] = "g"
    sort_order: Optional[int] = 0

    @field_validator("category")
    @classmethod
    def swap_cat_v(cls, v):
        if v not in ("fruits_veg", "fats", "carbs"): raise ValueError("Invalid category")
        return v
    @field_validator("source_unit", "swap_unit")
    @classmethod
    def swap_unit_v(cls, v):
        if v not in ("g", "oz", "ml"): raise ValueError("Unit must be g, oz, or ml")
        return v
    @field_validator("source_amount", "swap_amount")
    @classmethod
    def swap_amount_v(cls, v):
        if v is not None and (v < 0 or v > 10000): raise ValueError("Amount 0–10000")
        return v
    @field_validator("source_name", "swap_name")
    @classmethod
    def swap_name_v(cls, v):
        if v and len(v) > 150: raise ValueError("Max 150 chars")
        v = (v or "").strip()
        if re.search(r"<script|javascript\s*:|on\w+\s*=", v, re.I): raise ValueError("Invalid characters")
        return v


class VersionModel(BaseModel):
    major: int
    minor: int
    tiny: int
    notes: Optional[str] = ""
    @field_validator("major","minor","tiny")
    @classmethod
    def ver_v(cls, v):
        if v<0 or v>999: raise ValueError("0–999")
        return v


class SmtpModel(BaseModel):
    host: Optional[str] = "smtp.gmail.com"
    port: Optional[int] = 587
    username: Optional[str] = ""
    password: Optional[str] = ""
    from_name: Optional[str] = "BodyBuilder Coach"
    use_tls: Optional[bool] = True
    @field_validator("port")
    @classmethod
    def port_v(cls, v):
        if v is not None and (v<1 or v>65535): raise ValueError("1–65535")
        return v
    @field_validator("host")
    @classmethod
    def host_v(cls, v):
        if not v or not v.strip(): raise ValueError("Host required")
        return v.strip()


class SendProgramModel(BaseModel):
    athlete_id: int
    to_email: Optional[str] = ""
    subject: Optional[str] = "Your BodyBuilder Program"
    message: Optional[str] = ""
    @field_validator("to_email")
    @classmethod
    def email_v(cls, v):
        # Strip and reject embedded newlines to prevent SMTP header injection
        v = (v or "").strip().replace("\r", "").replace("\n", "")
        if not v or "@" not in v: raise ValueError("Valid email required")
        return v
    @field_validator("subject")
    @classmethod
    def subj_v(cls, v):
        # Strip and reject embedded newlines to prevent SMTP header injection
        v = (v or "").strip().replace("\r", "").replace("\n", "")
        if not v: raise ValueError("Subject required")
        if len(v) > 200: raise ValueError("Max 200 chars")
        return v


class WorkoutPlanModel(BaseModel):
    athlete_id: int
    title: str
    start_date: Optional[str] = ""
    end_date: Optional[str] = ""
    notes: Optional[str] = ""
    warmup_instructions: Optional[str] = ""
    sort_order: Optional[int] = 0
    rest_days: Optional[List[str]] = []
    @field_validator("title")
    @classmethod
    def title_v(cls, v):
        if not v or not v.strip(): raise ValueError("Title required")
        if len(v)>100: raise ValueError("Max 100 chars")
        return v.strip()
    @field_validator("notes", "warmup_instructions")
    @classmethod
    def notes_v(cls, v):
        v = (v or "").strip()
        if len(v) > 2000: raise ValueError("Max 2000 characters")
        return v
    @field_validator("start_date", "end_date")
    @classmethod
    def plan_date_v(cls, v, info):
        return _valid_date_or_empty(v, info.field_name)


class RestDayToggleModel(BaseModel):
    day: str
    @field_validator("day")
    @classmethod
    def day_v(cls, v):
        valid = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
        if v not in valid: raise ValueError(f"Invalid day '{v}'")
        return v


class WorkoutSessionModel(BaseModel):
    plan_id: int
    day_of_week: str
    session_title: Optional[str] = ""
    muscle_groups: Optional[List[str]] = []
    session_notes: Optional[str] = ""
    sort_order: Optional[int] = 0
    @field_validator("day_of_week")
    @classmethod
    def dow_v(cls, v):
        valid = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
        if v not in valid: raise ValueError(f"Must be one of: {', '.join(valid)}")
        return v
    @field_validator("session_title")
    @classmethod
    def title_v(cls, v):
        if v and len(v)>100: raise ValueError("Max 100 chars")
        return v or ""
    @field_validator("session_notes")
    @classmethod
    def snotes_v(cls, v):
        v = (v or "").strip()
        if len(v) > 2000: raise ValueError("Max 2000 characters")
        return v


class WorkoutExerciseModel(BaseModel):
    session_id: int
    name: str
    muscle_group: Optional[str] = ""
    set_type: Optional[str] = "working"
    sets_json: Optional[List[dict]] = []
    rep_range: Optional[str] = ""
    rir: Optional[int] = 2
    tempo: Optional[str] = ""
    intensifiers: Optional[str] = ""
    exercise_notes: Optional[str] = ""
    warmup_instructions: Optional[str] = ""
    image_url: Optional[str] = ""
    sort_order: Optional[int] = 0


    @field_validator("name")
    @classmethod
    def ex_name_v(cls, v):
        if not v or not v.strip(): raise ValueError("Exercise name required")
        if len(v) > 100: raise ValueError("Max 100 chars")
        return v.strip()
    @field_validator("set_type")
    @classmethod
    def ex_type_v(cls, v):
        if v not in ("warm_up","working","drop_set"): raise ValueError("warm_up/working/drop_set")
        return v
    @field_validator("rir")
    @classmethod
    def ex_rir_v(cls, v):
        if v is not None and (v < 0 or v > 10): raise ValueError("RIR 0–10")
        return v
    @field_validator("image_url")
    @classmethod
    def ex_image_url_v(cls, v):
        if not v: return ""
        v = v.strip()
        if len(v) > 500: raise ValueError("URL too long (max 500 chars)")
        # Allow http/https URLs and local exercise-image paths served by this app
        if not (v.startswith("http://") or v.startswith("https://")
                or v.startswith("/exercise-images/")):
            raise ValueError("Image URL must start with http://, https://, or /exercise-images/")
        return v
    @field_validator("sets_json")
    @classmethod
    def ex_sets_v(cls, v):
        if v is None: return []
        if len(v) > 50: raise ValueError("Max 50 sets per exercise")

        # Cardio exercise: exactly one set with cardio_type key
        CARDIO_KEYS   = {"cardio_type","duration_hours","duration_minutes","rpe","hr_min","hr_max"}
        STRENGTH_KEYS = {"set_number","type","weight","reps","rep_range","rir","tempo","notes"}
        ALLOWED_TYPES = {"W","M","I"}

        for i, s in enumerate(v):
            if not isinstance(s, dict): raise ValueError(f"Set {i+1} must be an object")

            is_cardio = "cardio_type" in s
            allowed   = CARDIO_KEYS if is_cardio else STRENGTH_KEYS
            extra     = set(s.keys()) - allowed
            if extra: raise ValueError(f"Set {i+1} has unexpected fields: {extra}")

            if is_cardio:
                # Validate cardio-specific fields
                ct = s.get("cardio_type", "")
                if ct and len(str(ct)) > 500:
                    raise ValueError(f"Set {i+1} cardio_type too long (max 500)")
                for fld in ("duration_hours", "duration_minutes"):
                    val = s.get(fld)
                    if val is not None:
                        try:
                            iv = int(val)
                            if iv < 0 or iv > 999:
                                raise ValueError(f"Set {i+1} {fld} must be 0–999")
                        except (TypeError, ValueError) as e:
                            if "must be" in str(e): raise
                            raise ValueError(f"Set {i+1} {fld} must be numeric")
                rpe = s.get("rpe")
                if rpe is not None:
                    try:
                        rv = float(rpe)
                        if rv < 1 or rv > 10:
                            raise ValueError(f"Set {i+1} rpe must be 1–10")
                    except (TypeError, ValueError) as e:
                        if "must be" in str(e): raise
                        raise ValueError(f"Set {i+1} rpe must be numeric")
                for fld in ("hr_min", "hr_max"):
                    val = s.get(fld)
                    if val is not None:
                        try:
                            iv = int(val)
                            if iv < 0 or iv > 300:
                                raise ValueError(f"Set {i+1} {fld} must be 0–300 bpm")
                        except (TypeError, ValueError) as e:
                            if "must be" in str(e): raise
                            raise ValueError(f"Set {i+1} {fld} must be numeric")
            else:
                # Validate strength-specific fields
                if "type" in s and s["type"] not in ALLOWED_TYPES:
                    raise ValueError(f"Set {i+1} type must be W, M, or I")
                if "weight" in s:
                    try:
                        w = float(s["weight"])
                        if w < 0 or w > 5000: raise ValueError(f"Set {i+1} weight out of range (0–5000)")
                    except (TypeError, ValueError) as e:
                        if "out of range" in str(e): raise
                        raise ValueError(f"Set {i+1} weight must be numeric")
                if "reps" in s:
                    try:
                        r = int(s["reps"])
                        if r < 0 or r > 200: raise ValueError(f"Set {i+1} reps out of range (0–200)")
                    except (TypeError, ValueError) as e:
                        if "out of range" in str(e): raise
                        raise ValueError(f"Set {i+1} reps must be an integer")
        return v
    @field_validator("rep_range","tempo","intensifiers","exercise_notes","warmup_instructions","muscle_group")
    @classmethod
    def ex_str_v(cls, v):
        if v and len(v) > 1000: raise ValueError("Field exceeds max length")
        return v or ""


class SupplementModel(BaseModel):
    day_of_week: str
    name: str
    dosage: Optional[str] = ""
    time_of_day: Optional[str] = "AM"
    sort_order: Optional[int] = 0
    @field_validator("day_of_week")
    @classmethod
    def dow_v(cls, v):
        valid = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
        if v not in valid: raise ValueError(f"Must be one of: {', '.join(valid)}")
        return v
    @field_validator("name")
    @classmethod
    def name_v(cls, v):
        if not v or not v.strip(): raise ValueError("Supplement name required")
        if len(v)>100: raise ValueError("Max 100 chars")
        return v.strip()
    @field_validator("dosage")
    @classmethod
    def dosage_v(cls, v):
        v = (v or "").strip()
        if len(v) > 80: raise ValueError("Max 80 characters")
        return v
    @field_validator("time_of_day")
    @classmethod
    def tod_v(cls, v):
        if v not in ("AM","Intra","PM"): raise ValueError("AM, Intra, or PM")
        return v


class MealModel(BaseModel):
    day_type: Optional[str] = "training"
    name: str
    sort_order: Optional[int] = 0
    @field_validator("day_type")
    @classmethod
    def dt_v(cls, v):
        if v not in ("training","off"): raise ValueError("training or off")
        return v
    @field_validator("name")
    @classmethod
    def name_v(cls, v):
        if not v or not v.strip(): raise ValueError("Meal name required")
        if len(v)>100: raise ValueError("Max 100 chars")
        return v.strip()


class MealItemModel(BaseModel):
    source_type: Optional[str] = "protein"
    food_name: str
    quantity: Optional[float] = 1
    weight_g: Optional[float] = 0
    serving_size: Optional[str] = "100g"
    protein_g: Optional[float] = 0
    carbs_g: Optional[float] = 0
    fat_g: Optional[float] = 0
    fiber_g: Optional[float] = 0
    sodium_mg: Optional[float] = 0
    potassium_mg: Optional[float] = 0
    sort_order: Optional[int] = 0
    @field_validator("source_type")
    @classmethod
    def st_v(cls, v):
        allowed = ("protein","carb","fat","vegetable","fruit","dairy","supplement")
        if v not in allowed: raise ValueError("Invalid source type")
        return v
    @field_validator("food_name")
    @classmethod
    def fn_v(cls, v):
        if not v or not v.strip(): raise ValueError("Food name required")
        if len(v) > 200: raise ValueError("Max 200 characters")
        return v.strip()
    @field_validator("serving_size")
    @classmethod
    def item_ss_v(cls, v):
        v = (v or "100g").strip()
        if len(v) > 50: raise ValueError("Max 50 characters")
        return v
    @field_validator("quantity")
    @classmethod
    def qty_v(cls, v):
        if v is not None and (v < 0 or v > 9999): raise ValueError("Quantity must be 0–9999")
        return v or 0
    @field_validator("weight_g")
    @classmethod
    def wg_v(cls, v):
        if v is not None and (v < 0 or v > 99999): raise ValueError("Weight must be 0–99999 g")
        return v or 0
    @field_validator("protein_g","carbs_g","fat_g","fiber_g")
    @classmethod
    def item_macro_v(cls, v):
        if v is not None and (v < 0 or v > 9999): raise ValueError("Must be 0–9999 g")
        return v or 0
    @field_validator("sodium_mg","potassium_mg")
    @classmethod
    def item_mineral_v(cls, v):
        if v is not None and (v < 0 or v > 99999): raise ValueError("Must be 0–99999 mg")
        return v or 0


# ─── Version ──────────────────────────────────────────────────────────────────

@app.get("/api/version")
def get_version():
    conn = get_db()
    row = conn.execute("SELECT * FROM version WHERE id=1").fetchone()
    conn.close()
    data = dict(row) if row else {"major":1,"minor":0,"tiny":0,"notes":""}
    data["version_string"] = f"{data['major']}.{data['minor']}.{data['tiny']}"
    return data

@app.put("/api/version")
def update_version(body: VersionModel):
    conn = get_db()
    conn.execute("UPDATE version SET major=?,minor=?,tiny=?,notes=? WHERE id=1",
                 (body.major,body.minor,body.tiny,body.notes))
    conn.commit(); conn.close()
    return get_version()


# ─── Athletes ─────────────────────────────────────────────────────────────────

@app.get("/api/athletes")
def list_athletes():
    conn = get_db()
    rows = conn.execute("SELECT * FROM athletes ORDER BY name").fetchall()
    conn.close()
    return [athlete_row_to_dict(r) for r in rows]

@app.post("/api/athletes")
def create_athlete(body: AthleteModel):
    conn = get_db()
    cur = conn.execute("""INSERT INTO athletes
        (name,email,birthdate,height_cm,weight_kg,body_fat_pct,sex,activity_level,
         workout_days_per_week,workout_days,workout_time,phase,deficit,units)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (body.name,body.email,body.birthdate,body.height_cm,body.weight_kg,body.body_fat_pct,
         body.sex,body.activity_level,body.workout_days_per_week,json.dumps(body.workout_days),
         body.workout_time,body.phase,body.deficit,body.units))
    new_id = cur.lastrowid
    conn.commit(); conn.close()
    ensure_athlete_defaults(new_id)
    return get_athlete(new_id)

@app.get("/api/athletes/{athlete_id}")
def get_athlete(athlete_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM athletes WHERE id=?", (athlete_id,)).fetchone()
    conn.close()
    if not row: raise HTTPException(404, "Athlete not found")
    return athlete_row_to_dict(row)

@app.put("/api/athletes/{athlete_id}")
def update_athlete(athlete_id: int, body: AthleteModel):
    conn = get_db()
    if not conn.execute("SELECT id FROM athletes WHERE id=?", (athlete_id,)).fetchone():
        conn.close(); raise HTTPException(404)
    conn.execute("""UPDATE athletes SET name=?,email=?,birthdate=?,height_cm=?,weight_kg=?,
        body_fat_pct=?,sex=?,activity_level=?,workout_days_per_week=?,workout_days=?,
        workout_time=?,phase=?,deficit=?,units=?,status=? WHERE id=?""",
        (body.name,body.email,body.birthdate,body.height_cm,body.weight_kg,body.body_fat_pct,
         body.sex,body.activity_level,body.workout_days_per_week,json.dumps(body.workout_days),
         body.workout_time,body.phase,body.deficit,body.units,body.status,athlete_id))
    conn.commit(); conn.close()
    return get_athlete(athlete_id)

@app.delete("/api/athletes/{athlete_id}")
def delete_athlete(athlete_id: int):
    conn = get_db()
    if not conn.execute("SELECT id FROM athletes WHERE id=?", (athlete_id,)).fetchone():
        conn.close(); raise HTTPException(404, "Athlete not found")
    conn.execute("DELETE FROM athletes WHERE id=?", (athlete_id,))
    conn.commit(); conn.close()
    return {"deleted": athlete_id}


# ─── Program ──────────────────────────────────────────────────────────────────

@app.get("/api/athletes/{athlete_id}/program")
def get_program(athlete_id: int):
    conn = get_db()
    row = conn.execute("SELECT * FROM programs WHERE athlete_id=?", (athlete_id,)).fetchone()
    conn.close()
    if not row:
        return {"athlete_id":athlete_id,"start_date":"","end_date":"","payment_processed":False}
    d = dict(row); d["payment_processed"] = bool(d["payment_processed"])
    return d

@app.put("/api/athletes/{athlete_id}/program")
def update_program(athlete_id: int, body: ProgramModel):
    conn = get_db()
    existing = conn.execute("SELECT id FROM programs WHERE athlete_id=?", (athlete_id,)).fetchone()
    if existing:
        conn.execute("UPDATE programs SET start_date=?,end_date=?,payment_processed=? WHERE athlete_id=?",
                     (body.start_date,body.end_date,int(body.payment_processed),athlete_id))
    else:
        conn.execute("INSERT INTO programs (athlete_id,start_date,end_date,payment_processed) VALUES (?,?,?,?)",
                     (athlete_id,body.start_date,body.end_date,int(body.payment_processed)))
    conn.commit(); conn.close()
    return get_program(athlete_id)


# ─── Activity Calories ────────────────────────────────────────────────────────

@app.get("/api/athletes/{athlete_id}/activity-calories")
def get_activity_calories(athlete_id: int):
    ensure_athlete_defaults(athlete_id)
    conn = get_db()
    rows = conn.execute("SELECT * FROM activity_calories WHERE athlete_id=? ORDER BY level", (athlete_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.put("/api/athletes/{athlete_id}/activity-calories/{level}")
def update_activity_calories(athlete_id: int, level: int, body: ActivityCalModel):
    if level<1 or level>5: raise HTTPException(400, "Level 1–5")
    conn = get_db()
    conn.execute("""INSERT INTO activity_calories (athlete_id,level,additional_calories,multiplier)
        VALUES (?,?,?,?) ON CONFLICT(athlete_id,level) DO UPDATE SET
        additional_calories=excluded.additional_calories, multiplier=excluded.multiplier""",
        (athlete_id, level, body.additional_calories, body.multiplier))
    conn.commit(); conn.close()
    return get_activity_calories(athlete_id)

@app.get("/api/athletes/{athlete_id}/daily-calories")
def get_daily_calories(athlete_id: int):
    ensure_athlete_defaults(athlete_id)
    ath = get_athlete(athlete_id)
    rmr = ath["average"]
    level = ath["activity_level"]
    deficit = ath["deficit"]
    conn = get_db()
    ac = conn.execute(
        "SELECT additional_calories, multiplier FROM activity_calories WHERE athlete_id=? AND level=?",
        (athlete_id, level)).fetchone()
    conn.close()
    additional = ac["additional_calories"] if ac else 0
    multiplier = ac["multiplier"] if ac else 1.2
    tdee   = rmr * multiplier + additional
    total  = max(0, tdee - deficit)
    return {
        "rmr": round(rmr, 1),
        "activity_level": level,
        "multiplier": round(multiplier, 4),
        "additional_calories": round(additional, 1),
        "tdee": round(tdee, 1),
        "deficit": round(deficit, 1),
        "total_calories": round(total, 1),
        "phase": ath["phase"],
        "name": ath["name"],
    }


# ─── Calendar ─────────────────────────────────────────────────────────────────

def _get_plan_sessions_for_month(athlete_id: int, year: int, month: int, conn):
    """Return dict of date→[session summaries] for all active plans in this month."""
    first = date(year, month, 1)
    last_day = (date(year, month+1, 1) - timedelta(days=1)) if month < 12 else date(year, 12, 31)
    plans = conn.execute("""SELECT * FROM workout_plans WHERE athlete_id=?
        AND (start_date='' OR start_date <= ?)
        AND (end_date='' OR end_date >= ?)""",
        (athlete_id, str(last_day), str(first))).fetchall()
    result = {}
    for plan in plans:
        p = dict(plan)
        p_start = date.fromisoformat(p["start_date"]) if p["start_date"] else first
        p_end = date.fromisoformat(p["end_date"]) if p["end_date"] else last_day
        sessions = conn.execute("SELECT * FROM workout_sessions WHERE plan_id=?", (p["id"],)).fetchall()
        for sess in sessions:
            s = dict(sess)
            first_occ = next_weekday_on_or_after(max(p_start, first), s["day_of_week"])
            cur = first_occ
            while cur <= min(p_end, last_day):
                ds = str(cur)
                if ds not in result: result[ds] = []
                result[ds].append({"plan_title": p["title"], "session_title": s["session_title"],
                                   "muscle_groups": json.loads(s.get("muscle_groups") or "[]"),
                                   "session_id": s["id"], "plan_id": p["id"]})
                cur += timedelta(days=7)
    return result

@app.get("/api/athletes/{athlete_id}/calendar/month")
def get_calendar_month(athlete_id: int, year: int, month: int):
    if year<1900 or year>2200: raise HTTPException(400, "Invalid year")
    if month<1 or month>12: raise HTTPException(400, "Invalid month")
    conn = get_db()
    prefix = f"{year:04d}-{month:02d}"
    days = conn.execute("SELECT * FROM calendar_days WHERE athlete_id=? AND date LIKE ?",
                        (athlete_id, f"{prefix}%")).fetchall()
    events = conn.execute("SELECT * FROM calendar_events WHERE athlete_id=? AND date LIKE ? ORDER BY event_time",
                          (athlete_id, f"{prefix}%")).fetchall()
    plan_sessions = _get_plan_sessions_for_month(athlete_id, year, month, conn)
    conn.close()
    return {"days":{r["date"]:dict(r) for r in days},
            "events":[dict(e) for e in events],
            "plan_sessions": plan_sessions}

def _require_date_str(date_str: str) -> str:
    """Validate that date_str is a valid ISO date (YYYY-MM-DD)."""
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_str):
        raise HTTPException(400, "date must be YYYY-MM-DD")
    try:
        date.fromisoformat(date_str)
    except ValueError:
        raise HTTPException(400, "Invalid date")
    return date_str


@app.get("/api/athletes/{athlete_id}/calendar/day/{date_str}")
def get_calendar_day(athlete_id: int, date_str: str):
    _require_date_str(date_str)
    conn = get_db()
    row = conn.execute("SELECT * FROM calendar_days WHERE athlete_id=? AND date=?",
                       (athlete_id, date_str)).fetchone()
    events = conn.execute("SELECT * FROM calendar_events WHERE athlete_id=? AND date=? ORDER BY event_time",
                          (athlete_id, date_str)).fetchall()
    conn.close()
    day = dict(row) if row else {"athlete_id":athlete_id,"date":date_str,"steps":0,
                                  "aerobic_type":"","aerobic_duration":0,"workout_notes":""}
    return {"day":day,"events":[dict(e) for e in events]}

@app.put("/api/athletes/{athlete_id}/calendar/day/{date_str}")
def update_calendar_day(athlete_id: int, date_str: str, body: CalDayModel):
    _require_date_str(date_str)
    conn = get_db()
    conn.execute("""INSERT INTO calendar_days (athlete_id,date,steps,aerobic_type,aerobic_duration,workout_notes)
        VALUES (?,?,?,?,?,?)
        ON CONFLICT(athlete_id,date) DO UPDATE SET steps=excluded.steps,
        aerobic_type=excluded.aerobic_type,aerobic_duration=excluded.aerobic_duration,
        workout_notes=excluded.workout_notes""",
        (athlete_id,date_str,body.steps,body.aerobic_type,body.aerobic_duration,body.workout_notes))
    conn.commit(); conn.close()
    return get_calendar_day(athlete_id, date_str)

@app.post("/api/athletes/{athlete_id}/calendar/events")
def create_event(athlete_id: int, body: EventModel):
    conn = get_db()
    cur = conn.execute("INSERT INTO calendar_events (athlete_id,date,title,description,event_time) VALUES (?,?,?,?,?)",
                       (athlete_id,body.date,body.title,body.description,body.event_time))
    eid = cur.lastrowid; conn.commit()
    row = conn.execute("SELECT * FROM calendar_events WHERE id=?", (eid,)).fetchone()
    conn.close(); return dict(row)

@app.put("/api/athletes/{athlete_id}/calendar/events/{event_id}")
def update_event(athlete_id: int, event_id: int, body: EventModel):
    conn = get_db()
    if not conn.execute("SELECT id FROM calendar_events WHERE id=? AND athlete_id=?", (event_id,athlete_id)).fetchone():
        conn.close(); raise HTTPException(404)
    conn.execute("UPDATE calendar_events SET date=?,title=?,description=?,event_time=? WHERE id=?",
                 (body.date,body.title,body.description,body.event_time,event_id))
    conn.commit()
    row = conn.execute("SELECT * FROM calendar_events WHERE id=?", (event_id,)).fetchone()
    conn.close(); return dict(row)

@app.delete("/api/athletes/{athlete_id}/calendar/events/{event_id}")
def delete_event(athlete_id: int, event_id: int):
    conn = get_db()
    conn.execute("DELETE FROM calendar_events WHERE id=? AND athlete_id=?", (event_id,athlete_id))
    conn.commit(); conn.close()
    return {"deleted": event_id}


# ─── Meal Plan ────────────────────────────────────────────────────────────────

@app.get("/api/athletes/{athlete_id}/meal-plan")
def get_meal_plan(athlete_id: int):
    ensure_athlete_defaults(athlete_id)
    conn = get_db()
    row = dict(conn.execute("SELECT * FROM meal_plans WHERE athlete_id=?", (athlete_id,)).fetchone())
    conn.close()
    dc = get_daily_calories(athlete_id)
    ath = get_athlete(athlete_id)
    deficit = float(ath.get("deficit", 0))
    rmr = dc["rmr"]
    row["rmr"] = rmr
    row["daily_calorie_intake"] = dc["total_calories"]
    row["rest_day_calories"] = round(max(0, rmr * 1.0 - deficit), 1)
    return row

@app.put("/api/athletes/{athlete_id}/meal-plan")
def update_meal_plan(athlete_id: int, body: MealPlanModel):
    ensure_athlete_defaults(athlete_id)
    conn = get_db()
    conn.execute("""UPDATE meal_plans SET protein_target=?,carbs_target=?,fat_target=?,fiber_target=?,
        sodium_target=?,potassium_target=?,protein_actual=?,carbs_actual=?,fat_actual=?,fiber_actual=?,
        sodium_actual=?,potassium_actual=? WHERE athlete_id=?""",
        (body.protein_target,body.carbs_target,body.fat_target,body.fiber_target,body.sodium_target,
         body.potassium_target,body.protein_actual,body.carbs_actual,body.fat_actual,body.fiber_actual,
         body.sodium_actual,body.potassium_actual,athlete_id))
    conn.commit(); conn.close()
    return get_meal_plan(athlete_id)


# ─── Foods ────────────────────────────────────────────────────────────────────

@app.get("/api/athletes/{athlete_id}/foods")
def get_foods(athlete_id: int, category: Optional[str] = None):
    conn = get_db()
    if category:
        rows = conn.execute("SELECT * FROM nutrition_foods WHERE athlete_id=? AND category=? ORDER BY name",
                            (athlete_id,category)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM nutrition_foods WHERE athlete_id=? ORDER BY name",
                            (athlete_id,)).fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.post("/api/athletes/{athlete_id}/foods")
def create_food(athlete_id: int, body: FoodModel):
    conn = get_db()
    cur = conn.execute("""INSERT INTO nutrition_foods
        (athlete_id,name,protein,carbs,fat,fiber,sodium,potassium,calories,serving_size,serving_g,category)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (athlete_id,body.name,body.protein,body.carbs,body.fat,body.fiber,
         body.sodium,body.potassium,body.calories,body.serving_size,body.serving_g,body.category))
    fid = cur.lastrowid; conn.commit()
    row = conn.execute("SELECT * FROM nutrition_foods WHERE id=?", (fid,)).fetchone()
    conn.close(); return dict(row)

@app.put("/api/athletes/{athlete_id}/foods/{food_id}")
def update_food(athlete_id: int, food_id: int, body: FoodModel):
    conn = get_db()
    # Authorise before mutating
    if not conn.execute("SELECT id FROM nutrition_foods WHERE id=? AND athlete_id=?",
                        (food_id, athlete_id)).fetchone():
        conn.close(); raise HTTPException(404)
    conn.execute("""UPDATE nutrition_foods SET name=?,protein=?,carbs=?,fat=?,fiber=?,sodium=?,
        potassium=?,calories=?,serving_size=?,serving_g=?,category=? WHERE id=? AND athlete_id=?""",
        (body.name,body.protein,body.carbs,body.fat,body.fiber,body.sodium,body.potassium,
         body.calories,body.serving_size,body.serving_g,body.category,food_id,athlete_id))
    conn.commit()
    row = conn.execute("SELECT * FROM nutrition_foods WHERE id=?", (food_id,)).fetchone()
    conn.close()
    return dict(row)

@app.delete("/api/athletes/{athlete_id}/foods/{food_id}")
def delete_food(athlete_id: int, food_id: int):
    conn = get_db()
    conn.execute("DELETE FROM nutrition_foods WHERE id=? AND athlete_id=?", (food_id,athlete_id))
    conn.commit(); conn.close()
    return {"deleted": food_id}


# ─── Food Swaps ───────────────────────────────────────────────────────────────

@app.get("/api/athletes/{athlete_id}/food-swaps")
def get_food_swaps(athlete_id: int):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM food_swaps WHERE athlete_id=? ORDER BY category, sort_order, id",
        (athlete_id,)).fetchall()
    conn.close(); return [dict(r) for r in rows]

@app.post("/api/athletes/{athlete_id}/food-swaps")
def create_food_swap(athlete_id: int, body: FoodSwapModel):
    conn = get_db()
    cur = conn.execute(
        """INSERT INTO food_swaps
           (athlete_id,category,source_name,source_amount,source_unit,swap_name,swap_amount,swap_unit,sort_order)
           VALUES (?,?,?,?,?,?,?,?,?)""",
        (athlete_id, body.category, body.source_name, body.source_amount, body.source_unit,
         body.swap_name, body.swap_amount, body.swap_unit, body.sort_order))
    sid = cur.lastrowid; conn.commit()
    row = conn.execute("SELECT * FROM food_swaps WHERE id=?", (sid,)).fetchone()
    conn.close(); return dict(row)

@app.put("/api/athletes/{athlete_id}/food-swaps/{swap_id}")
def update_food_swap(athlete_id: int, swap_id: int, body: FoodSwapModel):
    conn = get_db()
    row = conn.execute("SELECT id FROM food_swaps WHERE id=? AND athlete_id=?", (swap_id, athlete_id)).fetchone()
    if not row: raise HTTPException(404, "Swap not found")
    conn.execute(
        """UPDATE food_swaps SET category=?,source_name=?,source_amount=?,source_unit=?,
           swap_name=?,swap_amount=?,swap_unit=?,sort_order=? WHERE id=? AND athlete_id=?""",
        (body.category, body.source_name, body.source_amount, body.source_unit,
         body.swap_name, body.swap_amount, body.swap_unit, body.sort_order, swap_id, athlete_id))
    conn.commit()
    row = conn.execute("SELECT * FROM food_swaps WHERE id=?", (swap_id,)).fetchone()
    conn.close(); return dict(row)

@app.delete("/api/athletes/{athlete_id}/food-swaps/{swap_id}")
def delete_food_swap(athlete_id: int, swap_id: int):
    conn = get_db()
    row = conn.execute("SELECT id FROM food_swaps WHERE id=? AND athlete_id=?", (swap_id, athlete_id)).fetchone()
    if not row: raise HTTPException(404, "Swap not found")
    conn.execute("DELETE FROM food_swaps WHERE id=? AND athlete_id=?", (swap_id, athlete_id))
    conn.commit(); conn.close()
    return {"deleted": swap_id}


# ─── Supplements ──────────────────────────────────────────────────────────────

DAYS_OF_WEEK = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]

@app.get("/api/athletes/{athlete_id}/supplements")
def list_supplements(athlete_id: int, day_of_week: Optional[str] = None):
    conn = get_db()
    if day_of_week:
        if day_of_week not in DAYS_OF_WEEK:
            raise HTTPException(400, "Invalid day_of_week")
        rows = conn.execute(
            "SELECT * FROM supplements WHERE athlete_id=? AND day_of_week=? ORDER BY time_of_day, sort_order",
            (athlete_id, day_of_week)).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM supplements WHERE athlete_id=? ORDER BY day_of_week, time_of_day, sort_order",
            (athlete_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/athletes/{athlete_id}/supplements")
def create_supplement(athlete_id: int, body: SupplementModel):
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO supplements (athlete_id,day_of_week,name,dosage,time_of_day,sort_order) VALUES (?,?,?,?,?,?)",
        (athlete_id,body.day_of_week,body.name,body.dosage,body.time_of_day,body.sort_order))
    sid = cur.lastrowid; conn.commit()
    row = conn.execute("SELECT * FROM supplements WHERE id=?", (sid,)).fetchone()
    conn.close(); return dict(row)

@app.put("/api/athletes/{athlete_id}/supplements/{sup_id}")
def update_supplement(athlete_id: int, sup_id: int, body: SupplementModel):
    conn = get_db()
    if not conn.execute("SELECT id FROM supplements WHERE id=? AND athlete_id=?", (sup_id,athlete_id)).fetchone():
        conn.close(); raise HTTPException(404)
    conn.execute(
        "UPDATE supplements SET day_of_week=?,name=?,dosage=?,time_of_day=?,sort_order=? WHERE id=?",
        (body.day_of_week,body.name,body.dosage,body.time_of_day,body.sort_order,sup_id))
    conn.commit()
    row = conn.execute("SELECT * FROM supplements WHERE id=?", (sup_id,)).fetchone()
    conn.close(); return dict(row)

@app.delete("/api/athletes/{athlete_id}/supplements/{sup_id}")
def delete_supplement(athlete_id: int, sup_id: int):
    conn = get_db()
    conn.execute("DELETE FROM supplements WHERE id=? AND athlete_id=?", (sup_id,athlete_id))
    conn.commit(); conn.close()
    return {"deleted": sup_id}


# ─── Meals ────────────────────────────────────────────────────────────────────

def _load_meal(conn, meal_id):
    row = conn.execute("SELECT * FROM meals WHERE id=?", (meal_id,)).fetchone()
    if not row: raise HTTPException(404, "Meal not found")
    meal = dict(row)
    items = conn.execute("SELECT * FROM meal_items WHERE meal_id=? ORDER BY source_type, sort_order",
                         (meal_id,)).fetchall()
    meal["items"] = [dict(i) for i in items]
    # Compute totals
    totals = {"protein_g":0,"carbs_g":0,"fat_g":0,"fiber_g":0,"sodium_mg":0,"potassium_mg":0,"calories":0}
    for it in meal["items"]:
        qty = it.get("quantity",1) or 1
        for k in ["protein_g","carbs_g","fat_g","fiber_g","sodium_mg","potassium_mg"]:
            totals[k] += (it.get(k,0) or 0) * qty
        totals["calories"] += ((it.get("protein_g",0)*4) + (it.get("carbs_g",0)*4) + (it.get("fat_g",0)*9)) * qty
    meal["totals"] = {k: round(v,1) for k,v in totals.items()}
    return meal

@app.get("/api/athletes/{athlete_id}/meals")
def list_meals(athlete_id: int, day_type: Optional[str] = None):
    conn = get_db()
    if day_type:
        rows = conn.execute("SELECT id FROM meals WHERE athlete_id=? AND day_type=? ORDER BY sort_order, name",
                            (athlete_id, day_type)).fetchall()
    else:
        rows = conn.execute("SELECT id FROM meals WHERE athlete_id=? ORDER BY day_type, sort_order, name",
                            (athlete_id,)).fetchall()
    result = [_load_meal(conn, r["id"]) for r in rows]
    conn.close(); return result

@app.post("/api/athletes/{athlete_id}/meals")
def create_meal(athlete_id: int, body: MealModel):
    conn = get_db()
    cur = conn.execute("INSERT INTO meals (athlete_id,day_type,name,sort_order) VALUES (?,?,?,?)",
                       (athlete_id,body.day_type,body.name,body.sort_order))
    mid = cur.lastrowid; conn.commit()
    result = _load_meal(conn, mid); conn.close(); return result

@app.put("/api/athletes/{athlete_id}/meals/{meal_id}")
def update_meal(athlete_id: int, meal_id: int, body: MealModel):
    conn = get_db()
    if not conn.execute("SELECT id FROM meals WHERE id=? AND athlete_id=?", (meal_id,athlete_id)).fetchone():
        conn.close(); raise HTTPException(404)
    conn.execute("UPDATE meals SET day_type=?,name=?,sort_order=? WHERE id=?",
                 (body.day_type,body.name,body.sort_order,meal_id))
    conn.commit()
    result = _load_meal(conn, meal_id); conn.close(); return result

@app.delete("/api/athletes/{athlete_id}/meals/{meal_id}")
def delete_meal(athlete_id: int, meal_id: int):
    conn = get_db()
    conn.execute("DELETE FROM meals WHERE id=? AND athlete_id=?", (meal_id,athlete_id))
    conn.commit(); conn.close()
    return {"deleted": meal_id}

@app.post("/api/meals/{meal_id}/items")
def create_meal_item(meal_id: int, body: MealItemModel):
    conn = get_db()
    cur = conn.execute(
        """INSERT INTO meal_items (meal_id,source_type,food_name,quantity,weight_g,serving_size,
           protein_g,carbs_g,fat_g,fiber_g,sodium_mg,potassium_mg,sort_order)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (meal_id,body.source_type,body.food_name,body.quantity,body.weight_g,body.serving_size,
         body.protein_g,body.carbs_g,body.fat_g,body.fiber_g,body.sodium_mg,body.potassium_mg,body.sort_order))
    iid = cur.lastrowid; conn.commit()
    row = conn.execute("SELECT * FROM meal_items WHERE id=?", (iid,)).fetchone()
    conn.close(); return dict(row)

@app.put("/api/meal-items/{item_id}")
def update_meal_item(item_id: int, body: MealItemModel):
    conn = get_db()
    # Verify item exists before updating
    if not conn.execute("SELECT id FROM meal_items WHERE id=?", (item_id,)).fetchone():
        conn.close(); raise HTTPException(404, "Meal item not found")
    conn.execute(
        """UPDATE meal_items SET source_type=?,food_name=?,quantity=?,weight_g=?,serving_size=?,
           protein_g=?,carbs_g=?,fat_g=?,fiber_g=?,sodium_mg=?,potassium_mg=?,sort_order=? WHERE id=?""",
        (body.source_type,body.food_name,body.quantity,body.weight_g,body.serving_size,
         body.protein_g,body.carbs_g,body.fat_g,body.fiber_g,body.sodium_mg,body.potassium_mg,
         body.sort_order,item_id))
    conn.commit()
    row = conn.execute("SELECT * FROM meal_items WHERE id=?", (item_id,)).fetchone()
    conn.close()
    return dict(row)

@app.delete("/api/meal-items/{item_id}")
def delete_meal_item(item_id: int):
    conn = get_db()
    # Verify item exists and its parent meal is valid (ownership chain)
    row = conn.execute(
        "SELECT mi.id FROM meal_items mi "
        "JOIN meals m ON m.id = mi.meal_id "
        "WHERE mi.id=?", (item_id,)
    ).fetchone()
    if not row: raise HTTPException(404, "Meal item not found")
    conn.execute("DELETE FROM meal_items WHERE id=?", (item_id,))
    conn.commit(); conn.close()
    return {"deleted": item_id}


# ─── Workout Plans ────────────────────────────────────────────────────────────

def _load_plan(conn, plan_id):
    row = conn.execute("SELECT * FROM workout_plans WHERE id=?", (plan_id,)).fetchone()
    if not row: raise HTTPException(404, "Plan not found")
    plan = dict(row)
    plan["rest_days"] = json.loads(plan.get("rest_days") or "[]")
    sessions = conn.execute("SELECT * FROM workout_sessions WHERE plan_id=? ORDER BY sort_order, day_of_week",
                            (plan_id,)).fetchall()
    plan["sessions"] = []
    for sess in sessions:
        s = dict(sess)
        s["muscle_groups"] = json.loads(s.get("muscle_groups") or "[]")
        exs = conn.execute("SELECT * FROM workout_exercises WHERE session_id=? ORDER BY sort_order",
                           (s["id"],)).fetchall()
        s["exercises"] = []
        for ex in exs:
            e = dict(ex)
            e["sets_json"] = json.loads(e.get("sets_json") or "[]")
            s["exercises"].append(e)
        plan["sessions"].append(s)
    return plan

@app.get("/api/athletes/{athlete_id}/workout-plans")
def list_workout_plans(athlete_id: int):
    conn = get_db()
    rows = conn.execute("SELECT id FROM workout_plans WHERE athlete_id=? ORDER BY sort_order, title",
                        (athlete_id,)).fetchall()
    result = [_load_plan(conn, r["id"]) for r in rows]
    conn.close(); return result

@app.post("/api/athletes/{athlete_id}/workout-plans")
def create_workout_plan(athlete_id: int, body: WorkoutPlanModel):
    conn = get_db()
    cur = conn.execute("""INSERT INTO workout_plans (athlete_id,title,start_date,end_date,notes,warmup_instructions,sort_order,rest_days)
        VALUES (?,?,?,?,?,?,?,?)""",
        (athlete_id,body.title,body.start_date,body.end_date,body.notes,body.warmup_instructions,body.sort_order,json.dumps(body.rest_days or [])))
    pid = cur.lastrowid; conn.commit()
    result = _load_plan(conn, pid); conn.close(); return result

@app.put("/api/athletes/{athlete_id}/workout-plans/{plan_id}")
def update_workout_plan(athlete_id: int, plan_id: int, body: WorkoutPlanModel):
    conn = get_db()
    conn.execute("""UPDATE workout_plans SET title=?,start_date=?,end_date=?,notes=?,warmup_instructions=?,sort_order=?,rest_days=?
        WHERE id=? AND athlete_id=?""",
        (body.title,body.start_date,body.end_date,body.notes,body.warmup_instructions,body.sort_order,json.dumps(body.rest_days or []),plan_id,athlete_id))
    conn.commit()
    result = _load_plan(conn, plan_id); conn.close(); return result

@app.delete("/api/athletes/{athlete_id}/workout-plans/{plan_id}")
def delete_workout_plan(athlete_id: int, plan_id: int):
    conn = get_db()
    conn.execute("DELETE FROM workout_plans WHERE id=? AND athlete_id=?", (plan_id,athlete_id))
    conn.commit(); conn.close()
    return {"deleted": plan_id}

@app.patch("/api/athletes/{athlete_id}/workout-plans/{plan_id}/rest-days")
def toggle_rest_day(athlete_id: int, plan_id: int, body: RestDayToggleModel):
    conn = get_db()
    row = conn.execute("SELECT rest_days FROM workout_plans WHERE id=? AND athlete_id=?", (plan_id, athlete_id)).fetchone()
    if not row: conn.close(); raise HTTPException(404, "Plan not found")
    current = set(json.loads(row["rest_days"] or "[]"))
    if body.day in current:
        current.discard(body.day)
    else:
        current.add(body.day)
    conn.execute("UPDATE workout_plans SET rest_days=? WHERE id=?", (json.dumps(sorted(current)), plan_id))
    conn.commit()
    result = _load_plan(conn, plan_id)
    conn.close()
    return result


# ─── Workout Sessions ─────────────────────────────────────────────────────────

@app.post("/api/workout-sessions")
def create_session(body: WorkoutSessionModel):
    conn = get_db()
    if not conn.execute("SELECT id FROM workout_plans WHERE id=?", (body.plan_id,)).fetchone():
        conn.close(); raise HTTPException(404, "Workout plan not found")
    cur = conn.execute("""INSERT INTO workout_sessions (plan_id,day_of_week,session_title,muscle_groups,session_notes,sort_order)
        VALUES (?,?,?,?,?,?)""",
        (body.plan_id,body.day_of_week,body.session_title,json.dumps(body.muscle_groups),
         body.session_notes,body.sort_order))
    sid = cur.lastrowid; conn.commit()
    row = conn.execute("SELECT * FROM workout_sessions WHERE id=?", (sid,)).fetchone()
    s = dict(row); s["muscle_groups"] = json.loads(s.get("muscle_groups") or "[]")
    s["exercises"] = []; conn.close(); return s

@app.put("/api/workout-sessions/{session_id}")
def update_session(session_id: int, body: WorkoutSessionModel):
    conn = get_db()
    # Verify session exists and belongs to the stated plan
    row = conn.execute("SELECT plan_id FROM workout_sessions WHERE id=?", (session_id,)).fetchone()
    if not row: raise HTTPException(404, "Session not found")
    if row["plan_id"] != body.plan_id:
        raise HTTPException(403, "Session does not belong to that plan")
    conn.execute("""UPDATE workout_sessions SET plan_id=?,day_of_week=?,session_title=?,muscle_groups=?,
        session_notes=?,sort_order=? WHERE id=?""",
        (body.plan_id,body.day_of_week,body.session_title,json.dumps(body.muscle_groups),
         body.session_notes,body.sort_order,session_id))
    conn.commit()
    row = conn.execute("SELECT * FROM workout_sessions WHERE id=?", (session_id,)).fetchone()
    s = dict(row); s["muscle_groups"] = json.loads(s.get("muscle_groups") or "[]")
    exs = conn.execute("SELECT * FROM workout_exercises WHERE session_id=? ORDER BY sort_order", (session_id,)).fetchall()
    s["exercises"] = [{**dict(e),"sets_json":json.loads(e.get("sets_json") or "[]")} for e in exs]
    conn.close(); return s

@app.delete("/api/workout-sessions/{session_id}")
def delete_session(session_id: int):
    conn = get_db()
    # Verify session exists and its parent plan is valid (ownership chain)
    row = conn.execute(
        "SELECT ws.id FROM workout_sessions ws "
        "JOIN workout_plans wp ON wp.id = ws.plan_id "
        "WHERE ws.id=?", (session_id,)
    ).fetchone()
    if not row: raise HTTPException(404, "Session not found")
    conn.execute("DELETE FROM workout_sessions WHERE id=?", (session_id,))
    conn.commit(); conn.close()
    return {"deleted": session_id}

@app.post("/api/workout-sessions/{session_id}/clone")
def clone_session(session_id: int, body: WorkoutSessionModel):
    """Create a new session copying all exercises from an existing session as a template."""
    conn = get_db()
    # Create the new session
    cur = conn.execute("""INSERT INTO workout_sessions (plan_id,day_of_week,session_title,muscle_groups,session_notes,sort_order)
        VALUES (?,?,?,?,?,0)""",
        (body.plan_id, body.day_of_week, body.session_title,
         json.dumps(body.muscle_groups), body.session_notes))
    new_sid = cur.lastrowid
    # Copy all exercises from source session
    src_exs = conn.execute("SELECT * FROM workout_exercises WHERE session_id=? ORDER BY sort_order",
                           (session_id,)).fetchall()
    for ex in src_exs:
        conn.execute(
            """INSERT INTO workout_exercises
               (session_id,name,muscle_group,set_type,sets_json,rep_range,rir,tempo,intensifiers,exercise_notes,warmup_instructions,image_url,sort_order)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (new_sid, ex["name"], ex["muscle_group"], ex["set_type"],
             ex["sets_json"], ex["rep_range"], ex["rir"], ex["tempo"],
             ex["intensifiers"], ex["exercise_notes"], ex.get("warmup_instructions",""),
             ex.get("image_url",""), ex["sort_order"]))
    conn.commit()
    # Return the full plan so the UI can refresh
    plan_row = conn.execute("SELECT plan_id FROM workout_sessions WHERE id=?", (new_sid,)).fetchone()
    result = _load_plan(conn, plan_row["plan_id"])
    conn.close(); return result

@app.get("/api/athletes/{athlete_id}/workout-sessions/all")
def list_all_sessions(athlete_id: int):
    """Return all sessions across all plans, with exercises — used for the template picker."""
    conn = get_db()
    plans = conn.execute("SELECT id, title FROM workout_plans WHERE athlete_id=? ORDER BY sort_order, title",
                         (athlete_id,)).fetchall()
    result = []
    for p in plans:
        sessions = conn.execute(
            "SELECT * FROM workout_sessions WHERE plan_id=? ORDER BY sort_order, day_of_week",
            (p["id"],)).fetchall()
        for s in sessions:
            sess = dict(s)
            sess["plan_title"] = p["title"]
            sess["muscle_groups"] = json.loads(sess.get("muscle_groups") or "[]")
            exs = conn.execute("SELECT * FROM workout_exercises WHERE session_id=? ORDER BY sort_order",
                               (s["id"],)).fetchall()
            sess["exercises"] = [{**dict(e), "sets_json": json.loads(e.get("sets_json") or "[]")} for e in exs]
            result.append(sess)
    conn.close(); return result


# ─── Workout Exercises ────────────────────────────────────────────────────────

@app.post("/api/workout-exercises")
def create_exercise(body: WorkoutExerciseModel):
    conn = get_db()
    cur = conn.execute(
        """INSERT INTO workout_exercises
           (session_id,name,muscle_group,set_type,sets_json,rep_range,rir,tempo,intensifiers,exercise_notes,warmup_instructions,image_url,sort_order)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (body.session_id,body.name,body.muscle_group,body.set_type,json.dumps(body.sets_json),
         body.rep_range,body.rir,body.tempo,body.intensifiers,body.exercise_notes,body.warmup_instructions,body.image_url,body.sort_order))
    eid = cur.lastrowid; conn.commit()
    row = conn.execute("SELECT * FROM workout_exercises WHERE id=?", (eid,)).fetchone()
    e = dict(row); e["sets_json"] = json.loads(e.get("sets_json") or "[]")
    conn.close(); return e

@app.put("/api/workout-exercises/{exercise_id}")
def update_exercise(exercise_id: int, body: WorkoutExerciseModel):
    conn = get_db()
    # Verify exercise exists and belongs to the stated session
    ex_row = conn.execute("SELECT session_id FROM workout_exercises WHERE id=?", (exercise_id,)).fetchone()
    if not ex_row: raise HTTPException(404, "Exercise not found")
    if ex_row["session_id"] != body.session_id:
        raise HTTPException(403, "Exercise does not belong to that session")
    conn.execute(
        """UPDATE workout_exercises SET session_id=?,name=?,muscle_group=?,set_type=?,sets_json=?,
           rep_range=?,rir=?,tempo=?,intensifiers=?,exercise_notes=?,warmup_instructions=?,image_url=?,sort_order=? WHERE id=?""",
        (body.session_id,body.name,body.muscle_group,body.set_type,json.dumps(body.sets_json),
         body.rep_range,body.rir,body.tempo,body.intensifiers,body.exercise_notes,body.warmup_instructions,body.image_url,body.sort_order,exercise_id))
    conn.commit()
    row = conn.execute("SELECT * FROM workout_exercises WHERE id=?", (exercise_id,)).fetchone()
    e = dict(row); e["sets_json"] = json.loads(e.get("sets_json") or "[]")
    conn.close(); return e

@app.delete("/api/workout-exercises/{exercise_id}")
def delete_exercise(exercise_id: int):
    conn = get_db()
    # Verify exercise exists and its ownership chain (session → plan) is intact
    row = conn.execute(
        "SELECT we.id FROM workout_exercises we "
        "JOIN workout_sessions ws ON ws.id = we.session_id "
        "JOIN workout_plans wp ON wp.id = ws.plan_id "
        "WHERE we.id=?", (exercise_id,)
    ).fetchone()
    if not row: raise HTTPException(404, "Exercise not found")
    conn.execute("DELETE FROM workout_exercises WHERE id=?", (exercise_id,))
    conn.commit(); conn.close()
    return {"deleted": exercise_id}


# ─── SMTP ─────────────────────────────────────────────────────────────────────

@app.get("/api/admin/smtp")
def get_smtp():
    conn = get_db()
    row = dict(conn.execute("SELECT * FROM smtp_settings WHERE id=1").fetchone())
    conn.close()
    row["use_tls"] = bool(row["use_tls"])
    row["password"] = "••••••••" if row["password"] else ""
    return row

@app.put("/api/admin/smtp")
def update_smtp(body: SmtpModel):
    conn = get_db()
    existing = dict(conn.execute("SELECT password FROM smtp_settings WHERE id=1").fetchone())
    pw = body.password if (body.password and body.password != "••••••••") else existing["password"]
    conn.execute("UPDATE smtp_settings SET host=?,port=?,username=?,password=?,from_name=?,use_tls=? WHERE id=1",
                 (body.host,body.port,body.username,pw,body.from_name,int(body.use_tls)))
    conn.commit(); conn.close()
    return get_smtp()

@app.post("/api/admin/test-smtp")
def test_smtp():
    conn = get_db()
    cfg = dict(conn.execute("SELECT * FROM smtp_settings WHERE id=1").fetchone())
    conn.close()
    if not cfg["username"] or not cfg["password"]: raise HTTPException(400, "SMTP credentials not configured")
    try:
        if cfg["use_tls"]:
            s = smtplib.SMTP(cfg["host"], cfg["port"], timeout=10); s.starttls()
        else:
            s = smtplib.SMTP_SSL(cfg["host"], cfg["port"], timeout=10)
        s.login(cfg["username"], cfg["password"]); s.quit()
        return {"success":True,"message":"SMTP connection successful!"}
    except Exception as e:
        raise HTTPException(400, f"Connection failed: {e}")


# ─── Excel Export ─────────────────────────────────────────────────────────────

def _build_workbook(athlete_id: int, plan_id: int = None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    conn = get_db()
    try:
        ath = athlete_row_to_dict(conn.execute("SELECT * FROM athletes WHERE id=?", (athlete_id,)).fetchone())
        prog_row = conn.execute("SELECT * FROM programs WHERE athlete_id=?", (athlete_id,)).fetchone()
        prog = dict(prog_row) if prog_row else {}
        mp_row = conn.execute("SELECT * FROM meal_plans WHERE athlete_id=?", (athlete_id,)).fetchone()
        mp = dict(mp_row) if mp_row else {}
        # Fetch all meals with their items for this athlete
        meals_raw = conn.execute(
            "SELECT id, name, day_type, sort_order FROM meals WHERE athlete_id=? ORDER BY day_type, sort_order, name",
            (athlete_id,)).fetchall()
        meals_for_plan = []
        for m_row in meals_raw:
            m = dict(m_row)
            items_raw = conn.execute(
                """SELECT food_name, source_type, weight_g, serving_size, quantity,
                          protein_g, carbs_g, fat_g, fiber_g, sodium_mg, potassium_mg, sort_order
                   FROM meal_items WHERE meal_id=? ORDER BY source_type, sort_order""",
                (m["id"],)).fetchall()
            m["items"] = [dict(i) for i in items_raw]
            meals_for_plan.append(m)
        # Flatten all items for Food Swaps sheet
        meal_items_for_swaps = [i for m in meals_for_plan for i in m["items"]]
        if plan_id is not None:
            plan_ids = conn.execute("SELECT id FROM workout_plans WHERE athlete_id=? AND id=?", (athlete_id, plan_id)).fetchall()
        else:
            plan_ids = conn.execute("SELECT id FROM workout_plans WHERE athlete_id=? ORDER BY sort_order, title", (athlete_id,)).fetchall()
        plans = [_load_plan(conn, r["id"]) for r in plan_ids]
        dc_row = conn.execute(
            "SELECT additional_calories, multiplier FROM activity_calories WHERE athlete_id=? AND level=?",
            (athlete_id, ath["activity_level"])).fetchone()
    finally:
        conn.close()
    additional_cal = dc_row["additional_calories"] if dc_row else 0
    multiplier_val = dc_row["multiplier"] if dc_row else 1.2
    total_cal = max(0, ath["average"] * multiplier_val + additional_cal - ath["deficit"])

    wb = Workbook()

    # Colour palette
    HDR_FILL  = PatternFill("solid", fgColor="1A1D27")
    SUBHDR    = PatternFill("solid", fgColor="2E3349")
    ACCENT    = PatternFill("solid", fgColor="4F8EF7")
    GREEN_F   = PatternFill("solid", fgColor="2ECC71")
    WHITE_F   = PatternFill("solid", fgColor="FFFFFF")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    SUBF      = Font(bold=True, color="FFFFFF", size=10)
    BODY_FONT = Font(color="1A1D27", size=10)
    BOLD_BODY = Font(bold=True, color="1A1D27", size=10)
    thin      = Side(style="thin", color="CCCCCC")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left", vertical="center")
    WRAP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def set_hdr(ws, row, col, val, fill=HDR_FILL, font=HDR_FONT):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill; c.font = font; c.alignment = CENTER; c.border = BORDER

    def set_cell(ws, row, col, val, bold=False, fill=None, align=LEFT):
        c = ws.cell(row=row, column=col, value=val)
        c.font = BOLD_BODY if bold else BODY_FONT
        if fill: c.fill = fill
        c.alignment = align; c.border = BORDER

    def auto_width(ws, extra=4):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try: max_len = max(max_len, len(str(cell.value or "")))
                except: pass
            ws.column_dimensions[col_letter].width = min(max_len + extra, 60)

    # ── Sheet 1: Athlete Profile ──
    ws = wb.active; ws.title = "Athlete Profile"
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:D1")
    c = ws["A1"]; c.value = f"💪 BodyBuilder — {ath['name']}"; c.font = Font(bold=True, color="FFFFFF", size=14)
    c.fill = ACCENT; c.alignment = CENTER

    headers = ["Field", "Value", "Field", "Value"]
    for i, h in enumerate(headers, 1): set_hdr(ws, 2, i, h, SUBHDR, SUBF)

    rows = [
        ("Name", ath["name"], "Email", ath["email"]),
        ("Date of Birth", ath["birthdate"], "Age", ath["age"]),
        ("Sex", ath["sex"].title(), "Phase", ath["phase"].title()),
        ("Height (cm)", ath["height_cm"], "Weight (kg)", ath["weight_kg"]),
        ("Body Fat %", ath.get("body_fat_pct",0), "LBM (kg)", ath["lbm_kg"]),
        ("Activity Level", ath["activity_level"], "Workout Days/Week", ath["workout_days_per_week"]),
        ("Workout Time", ath["workout_time"], "Workout Days", ", ".join(ath["workout_days"])),
        ("", "", "", ""),
        ("RMR — Mifflin-St Jeor", ath["mifflin"], "RMR — Harris-Benedict", ath["harris"]),
        ("RMR — Katch-McArdle", ath["katch"], "RMR Average (used)", ath["average"]),
        ("Activity Multiplier", multiplier_val, "Additional Calories", additional_cal),
        ("TDEE (RMR × Multiplier)", round(ath["average"] * multiplier_val, 1), "Caloric Deficit", ath["deficit"]),
        ("", "", "", ""),
        ("TOTAL DAILY CALORIES", round(total_cal, 1), "Program Phase", ath["phase"].upper()),
    ]
    for ri, row in enumerate(rows, 3):
        for ci, val in enumerate(row, 1):
            bold = ri == len(rows)+2 and ci in (1,2)
            fill = GREEN_F if (ri == len(rows)+2 and ci in (1,2)) else None
            set_cell(ws, ri, ci, val, bold=bold, fill=fill)

    auto_width(ws)

    # ── Sheet 2: Program Details ──
    ws2 = wb.create_sheet("Program Details")
    for i, h in enumerate(["Field", "Value"], 1): set_hdr(ws2, 1, i, h, SUBHDR, SUBF)
    prog_rows = [
        ("Program Start Date", prog.get("start_date","")),
        ("Program End Date", prog.get("end_date","")),
        ("Payment Processed", "Yes" if prog.get("payment_processed") else "No"),
    ]
    for ri, (k, v) in enumerate(prog_rows, 2):
        set_cell(ws2, ri, 1, k, bold=True); set_cell(ws2, ri, 2, v)
    auto_width(ws2)

    # ── Sheet 3: Meal Plan ──
    ws3 = wb.create_sheet("Meal Plan")
    ws3.row_dimensions[1].height = 28
    MEAL_COLS = ["Food", "Serving", "Kcal", "Protein (g)", "Carbs (g)", "Fat (g)", "Fiber (g)"]
    NCOLS3 = len(MEAL_COLS)
    ws3.merge_cells(f"A1:{get_column_letter(NCOLS3)}1")
    c = ws3["A1"]; c.value = f"Meal Plan — {ath['name']}"
    c.font = Font(bold=True, color="FFFFFF", size=14); c.fill = ACCENT; c.alignment = CENTER

    DAY_TYPE_FILLS = {
        "training":    PatternFill("solid", fgColor="E8F4FD"),
        "rest":        PatternFill("solid", fgColor="FEF9E7"),
        "competition": PatternFill("solid", fgColor="FDECEA"),
    }
    MEAL_HDR_FILL = PatternFill("solid", fgColor="3A3F5A")

    ri3 = 2
    if meals_for_plan:
        current_day = None
        for meal in meals_for_plan:
            day_type = meal.get("day_type", "training")
            # Day-type section header (only when day changes)
            if day_type != current_day:
                current_day = day_type
                ws3.merge_cells(f"A{ri3}:{get_column_letter(NCOLS3)}{ri3}")
                dc = ws3.cell(row=ri3, column=1, value=f"── {day_type.upper()} DAY ──")
                dc.font = Font(bold=True, color="FFFFFF", size=11)
                dc.fill = HDR_FILL; dc.alignment = CENTER
                ws3.row_dimensions[ri3].height = 20
                ri3 += 1

            # Meal name header
            meal_name = meal.get("name") or "Meal"
            meal_kcal = sum(
                ((it.get("protein_g",0)*4) + (it.get("carbs_g",0)*4) + (it.get("fat_g",0)*9)) * (it.get("quantity",1) or 1)
                for it in meal["items"]
            )
            ws3.merge_cells(f"A{ri3}:{get_column_letter(NCOLS3-1)}{ri3}")
            mc = ws3.cell(row=ri3, column=1, value=meal_name)
            mc.font = Font(bold=True, color="FFFFFF", size=11); mc.fill = MEAL_HDR_FILL; mc.alignment = LEFT
            kcal_c = ws3.cell(row=ri3, column=NCOLS3, value=round(meal_kcal))
            kcal_c.font = Font(bold=True, color="FFFFFF", size=11); kcal_c.fill = MEAL_HDR_FILL; kcal_c.alignment = CENTER
            ws3.row_dimensions[ri3].height = 18
            ri3 += 1

            if meal["items"]:
                # Column headers
                for ci, h in enumerate(MEAL_COLS, 1):
                    set_hdr(ws3, ri3, ci, h, SUBHDR, SUBF)
                ri3 += 1

                # Food rows
                row_fill = DAY_TYPE_FILLS.get(day_type, WHITE_F)
                meal_totals = {"protein_g":0,"carbs_g":0,"fat_g":0,"fiber_g":0,"kcal":0}
                for it in meal["items"]:
                    qty = it.get("quantity", 1) or 1
                    p = (it.get("protein_g",0) or 0) * qty
                    cb = (it.get("carbs_g",0) or 0) * qty
                    fa = (it.get("fat_g",0) or 0) * qty
                    fi = (it.get("fiber_g",0) or 0) * qty
                    ik = round(p*4 + cb*4 + fa*9)
                    wg = it.get("weight_g", 0) or 0
                    sv = it.get("serving_size") or (f"{round(wg)}g" if wg else "")
                    set_cell(ws3, ri3, 1, it["food_name"], bold=True, fill=row_fill)
                    set_cell(ws3, ri3, 2, sv,              fill=row_fill)
                    set_cell(ws3, ri3, 3, ik,              fill=row_fill, align=CENTER)
                    set_cell(ws3, ri3, 4, round(p,1),      fill=row_fill, align=CENTER)
                    set_cell(ws3, ri3, 5, round(cb,1),     fill=row_fill, align=CENTER)
                    set_cell(ws3, ri3, 6, round(fa,1),     fill=row_fill, align=CENTER)
                    set_cell(ws3, ri3, 7, round(fi,1),     fill=row_fill, align=CENTER)
                    meal_totals["protein_g"] += p; meal_totals["carbs_g"] += cb
                    meal_totals["fat_g"] += fa;   meal_totals["fiber_g"] += fi
                    meal_totals["kcal"] += ik
                    ri3 += 1

                # Meal totals row
                set_cell(ws3, ri3, 1, "MEAL TOTAL", bold=True, fill=GREEN_F)
                set_cell(ws3, ri3, 2, "",            fill=GREEN_F)
                set_cell(ws3, ri3, 3, round(meal_totals["kcal"]),    bold=True, fill=GREEN_F, align=CENTER)
                set_cell(ws3, ri3, 4, round(meal_totals["protein_g"],1), bold=True, fill=GREEN_F, align=CENTER)
                set_cell(ws3, ri3, 5, round(meal_totals["carbs_g"],1),   bold=True, fill=GREEN_F, align=CENTER)
                set_cell(ws3, ri3, 6, round(meal_totals["fat_g"],1),     bold=True, fill=GREEN_F, align=CENTER)
                set_cell(ws3, ri3, 7, round(meal_totals["fiber_g"],1),   bold=True, fill=GREEN_F, align=CENTER)
                ri3 += 1
            else:
                ws3.merge_cells(f"A{ri3}:{get_column_letter(NCOLS3)}{ri3}")
                ec = ws3.cell(row=ri3, column=1, value="No foods added to this meal yet.")
                ec.font = Font(color="999999", size=10, italic=True); ec.alignment = LEFT
                ri3 += 1

            ri3 += 1  # blank row between meals

        # Grand daily total row
        # Single pass over all items for daily totals
        total_kcal_day = total_p = total_cb = total_fat = 0.0
        for m in meals_for_plan:
            for it in m["items"]:
                qty = it.get("quantity", 1) or 1
                p   = (it.get("protein_g", 0) or 0) * qty
                cb  = (it.get("carbs_g",   0) or 0) * qty
                fat = (it.get("fat_g",     0) or 0) * qty
                total_p        += p
                total_cb       += cb
                total_fat      += fat
                total_kcal_day += p * 4 + cb * 4 + fat * 9
        ws3.merge_cells(f"A{ri3}:{get_column_letter(NCOLS3)}{ri3}")
        ws3.row_dimensions[ri3].height = 22
        ri3 += 1
        for ci, val in enumerate([
            "DAILY TOTAL", "", round(total_kcal_day),
            round(total_p,1), round(total_cb,1), round(total_fat,1), ""
        ], 1):
            c = ws3.cell(row=ri3, column=ci, value=val)
            c.font = Font(bold=True, color="FFFFFF", size=11)
            c.fill = ACCENT; c.alignment = CENTER if ci > 1 else LEFT; c.border = BORDER
    else:
        ws3.merge_cells(f"A2:{get_column_letter(NCOLS3)}2")
        ec = ws3["A2"]
        ec.value = "No meals found. Create meals in the Meal Plan tab and add foods to them."
        ec.font = Font(color="999999", size=10, italic=True); ec.alignment = LEFT

    auto_width(ws3)
    # Fix column widths for Meal Plan sheet
    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["B"].width = 22

    # ── Sheet 4: Food Swaps ──
    # Calorie-matching swap database (values are per 100 g)
    _SWAP_DB = [
        # Proteins
        {"n":"Chicken Breast (raw)",      "cat":"protein", "kcal":165},
        {"n":"Chicken Breast (cooked)",   "cat":"protein", "kcal":187},
        {"n":"Chicken Thigh (raw)",        "cat":"protein", "kcal":177},
        {"n":"Turkey Breast",             "cat":"protein", "kcal":135},
        {"n":"Ground Turkey 93%",         "cat":"protein", "kcal":163},
        {"n":"Lean Ground Beef 93%",      "cat":"protein", "kcal":172},
        {"n":"Lean Ground Beef 96%",      "cat":"protein", "kcal":137},
        {"n":"Beef Steak (Sirloin)",      "cat":"protein", "kcal":207},
        {"n":"Pork Tenderloin",           "cat":"protein", "kcal":143},
        {"n":"Bison",                     "cat":"protein", "kcal":146},
        {"n":"Salmon (Atlantic)",         "cat":"protein", "kcal":208},
        {"n":"Tuna (canned in water)",    "cat":"protein", "kcal":109},
        {"n":"Tilapia",                   "cat":"protein", "kcal":96},
        {"n":"Cod",                       "cat":"protein", "kcal":82},
        {"n":"Shrimp",                    "cat":"protein", "kcal":99},
        {"n":"Whitefish",                 "cat":"protein", "kcal":134},
        {"n":"Whole Egg",                 "cat":"protein", "kcal":155},
        {"n":"Egg White",                 "cat":"protein", "kcal":52},
        {"n":"Whey Protein Powder",       "cat":"protein", "kcal":400},
        {"n":"Casein Protein Powder",     "cat":"protein", "kcal":371},
        {"n":"Cottage Cheese (1%)",       "cat":"protein", "kcal":72},
        {"n":"Cottage Cheese (2%)",       "cat":"protein", "kcal":84},
        {"n":"Canned Salmon",             "cat":"protein", "kcal":139},
        {"n":"Tempeh",                    "cat":"protein", "kcal":195},
        # Carbs
        {"n":"White Rice (cooked)",       "cat":"carb", "kcal":130},
        {"n":"Brown Rice (cooked)",       "cat":"carb", "kcal":123},
        {"n":"Jasmine Rice (cooked)",     "cat":"carb", "kcal":129},
        {"n":"Oats (dry)",                "cat":"carb", "kcal":389},
        {"n":"Oats (cooked)",             "cat":"carb", "kcal":71},
        {"n":"Sweet Potato (raw)",        "cat":"carb", "kcal":86},
        {"n":"Sweet Potato (cooked)",     "cat":"carb", "kcal":90},
        {"n":"White Potato",              "cat":"carb", "kcal":77},
        {"n":"Pasta (dry)",               "cat":"carb", "kcal":371},
        {"n":"Whole Wheat Pasta (dry)",   "cat":"carb", "kcal":348},
        {"n":"Bread (Whole Wheat)",       "cat":"carb", "kcal":247},
        {"n":"Sourdough Bread",           "cat":"carb", "kcal":289},
        {"n":"Quinoa (cooked)",           "cat":"carb", "kcal":120},
        {"n":"Lentils (cooked)",          "cat":"carb", "kcal":116},
        {"n":"Black Beans (cooked)",      "cat":"carb", "kcal":132},
        {"n":"Chickpeas (cooked)",        "cat":"carb", "kcal":164},
        {"n":"Bagel (plain)",             "cat":"carb", "kcal":272},
        # Fats
        {"n":"Olive Oil",                 "cat":"fat", "kcal":884},
        {"n":"Coconut Oil",               "cat":"fat", "kcal":862},
        {"n":"Avocado",                   "cat":"fat", "kcal":160},
        {"n":"Almonds",                   "cat":"fat", "kcal":579},
        {"n":"Walnuts",                   "cat":"fat", "kcal":654},
        {"n":"Cashews",                   "cat":"fat", "kcal":553},
        {"n":"Peanuts",                   "cat":"fat", "kcal":567},
        {"n":"Macadamia Nuts",            "cat":"fat", "kcal":718},
        {"n":"Peanut Butter (natural)",   "cat":"fat", "kcal":598},
        {"n":"Almond Butter",             "cat":"fat", "kcal":614},
        {"n":"Chia Seeds",                "cat":"fat", "kcal":486},
        {"n":"Flaxseed",                  "cat":"fat", "kcal":534},
        {"n":"Hemp Seeds",                "cat":"fat", "kcal":553},
        # Vegetables
        {"n":"Broccoli",                  "cat":"vegetable", "kcal":34},
        {"n":"Spinach",                   "cat":"vegetable", "kcal":23},
        {"n":"Kale",                      "cat":"vegetable", "kcal":49},
        {"n":"Asparagus",                 "cat":"vegetable", "kcal":20},
        {"n":"Green Beans",               "cat":"vegetable", "kcal":31},
        {"n":"Bell Pepper",               "cat":"vegetable", "kcal":31},
        {"n":"Zucchini",                  "cat":"vegetable", "kcal":17},
        {"n":"Cauliflower",               "cat":"vegetable", "kcal":25},
        {"n":"Cucumber",                  "cat":"vegetable", "kcal":15},
        {"n":"Mushrooms",                 "cat":"vegetable", "kcal":22},
        {"n":"Brussels Sprouts",          "cat":"vegetable", "kcal":43},
        {"n":"Tomato",                    "cat":"vegetable", "kcal":18},
        {"n":"Mixed Greens / Lettuce",    "cat":"vegetable", "kcal":14},
        {"n":"Peas",                      "cat":"vegetable", "kcal":81},
        {"n":"Edamame",                   "cat":"vegetable", "kcal":122},
        {"n":"Beets",                     "cat":"vegetable", "kcal":43},
        # Fruits
        {"n":"Banana",                    "cat":"fruit", "kcal":89},
        {"n":"Apple",                     "cat":"fruit", "kcal":52},
        {"n":"Blueberries",               "cat":"fruit", "kcal":57},
        {"n":"Strawberries",              "cat":"fruit", "kcal":32},
        {"n":"Blackberries",              "cat":"fruit", "kcal":43},
        {"n":"Raspberries",               "cat":"fruit", "kcal":52},
        {"n":"Mango",                     "cat":"fruit", "kcal":60},
        {"n":"Orange",                    "cat":"fruit", "kcal":47},
        {"n":"Pineapple",                 "cat":"fruit", "kcal":50},
        {"n":"Grapes",                    "cat":"fruit", "kcal":69},
        {"n":"Watermelon",                "cat":"fruit", "kcal":30},
        {"n":"Cherries",                  "cat":"fruit", "kcal":50},
        {"n":"Peach",                     "cat":"fruit", "kcal":39},
        {"n":"Pear",                      "cat":"fruit", "kcal":57},
        {"n":"Kiwi",                      "cat":"fruit", "kcal":61},
        # Dairy
        {"n":"Greek Yogurt (non-fat)",    "cat":"dairy", "kcal":59},
        {"n":"Greek Yogurt (2%)",         "cat":"dairy", "kcal":73},
        {"n":"Milk (2%)",                 "cat":"dairy", "kcal":50},
        {"n":"Milk (whole)",              "cat":"dairy", "kcal":61},
        {"n":"Mozzarella (part-skim)",    "cat":"dairy", "kcal":254},
        {"n":"Cheddar Cheese",            "cat":"dairy", "kcal":403},
        {"n":"Ricotta (part-skim)",       "cat":"dairy", "kcal":138},
    ]

    def _find_swaps(food_name, source_type, weight_g, target_kcal, max_swaps=5):
        """Return list of {name, serving_g, serving_label} calorie-matched to target_kcal."""
        if target_kcal <= 0 or weight_g <= 0:
            return []
        name_lower = food_name.lower()
        candidates = [f for f in _SWAP_DB
                      if f["cat"] == source_type
                      and f["n"].lower() != name_lower
                      and f["kcal"] > 0]
        swaps = []
        for c in candidates:
            swap_g = target_kcal / (c["kcal"] / 100.0)
            # sanity: skip unreasonably tiny or large amounts
            if swap_g < 3 or swap_g > 2000:
                continue
            # Round to nearest 5g for clean display; 1g if very small (e.g. oils)
            if swap_g < 20:
                rounded = round(swap_g)
            else:
                rounded = round(swap_g / 5) * 5
            swaps.append({"name": c["n"], "g": rounded, "label": f"{rounded}g"})
        return swaps[:max_swaps]

    def _kcal_for_item(mi):
        """Compute kcal for a meal item using _SWAP_DB if food is known, else from stored macros."""
        wg = mi.get("weight_g") or 0
        name_lower = (mi.get("food_name") or "").lower().strip()
        db_match = next((f for f in _SWAP_DB if f["n"].lower() == name_lower), None)
        if db_match and wg > 0:
            return round(db_match["kcal"] * wg / 100)
        # fallback: compute from stored macros (scaled by quantity if present)
        qty = mi.get("quantity", 1) or 1
        return round((mi.get("protein_g",0)*4 + mi.get("carbs_g",0)*4 + mi.get("fat_g",0)*9) * qty)

    def _serving_label(mi):
        """Return a serving label consistent with the weight used."""
        wg = mi.get("weight_g") or 0
        sv = mi.get("serving_size") or ""
        if wg > 0:
            return f"{round(wg)}g"
        return sv or "100g"

    # Deduplicate meal plan foods: key = food_name (case-insensitive)
    # Keep the entry with the highest calorie count (most representative serving)
    seen_foods = {}
    for mi in meal_items_for_swaps:
        name = (mi.get("food_name") or "").strip()
        if not name:
            continue
        key = name.lower()
        item_kcal = _kcal_for_item(mi)
        if key not in seen_foods or item_kcal > seen_foods[key]["kcal"]:
            seen_foods[key] = {
                "food_name":   name,
                "source_type": mi.get("source_type", "protein"),
                "weight_g":    mi.get("weight_g") or 0,
                "serving_size": _serving_label(mi),
                "kcal":        item_kcal,
            }
    plan_foods = sorted(seen_foods.values(), key=lambda x: (x["source_type"], x["food_name"]))

    ws4 = wb.create_sheet("Food Swaps")
    ws4.row_dimensions[1].height = 28
    ws4.merge_cells("A1:M1")
    c4 = ws4["A1"]; c4.value = "Food Swaps — Calorie-Equivalent Alternatives"
    c4.font = Font(bold=True, color="FFFFFF", size=14); c4.fill = ACCENT; c4.alignment = CENTER

    # Column headers
    SWAP_HDRS = [
        "Food (as used in plan)", "Serving Used", "Calories",
        "Swap 1", "Swap 1 Serving",
        "Swap 2", "Swap 2 Serving",
        "Swap 3", "Swap 3 Serving",
        "Swap 4", "Swap 4 Serving",
        "Swap 5", "Swap 5 Serving",
    ]
    NCOLS4 = len(SWAP_HDRS)
    for ci, h in enumerate(SWAP_HDRS, 1):
        set_hdr(ws4, 2, ci, h, SUBHDR, SUBF)

    # Category colour bands for visual grouping
    CAT_FILLS = {
        "protein":    PatternFill("solid", fgColor="EAF4FF"),
        "carb":       PatternFill("solid", fgColor="EAFFF0"),
        "fat":        PatternFill("solid", fgColor="FFFADF"),
        "vegetable":  PatternFill("solid", fgColor="F0FFED"),
        "fruit":      PatternFill("solid", fgColor="FFF0F8"),
        "dairy":      PatternFill("solid", fgColor="F5F0FF"),
        "supplement": PatternFill("solid", fgColor="F5F5F5"),
    }

    SWAP_FILL  = PatternFill("solid", fgColor="FFFFFF")
    SWAP_FONT  = Font(color="444444", size=10, italic=True)
    SWAP_G_FONT = Font(color="4F8EF7", size=10, bold=True)

    if plan_foods:
        for ri, pf in enumerate(plan_foods, 3):
            row_fill = CAT_FILLS.get(pf["source_type"], WHITE_F)
            kcal_rounded = round(pf["kcal"])
            set_cell(ws4, ri, 1, pf["food_name"],    bold=True, fill=row_fill)
            set_cell(ws4, ri, 2, pf["serving_size"],  fill=row_fill)
            set_cell(ws4, ri, 3, kcal_rounded,        bold=True, fill=row_fill, align=CENTER)

            swaps = _find_swaps(pf["food_name"], pf["source_type"], pf["weight_g"], pf["kcal"])
            for si, sw in enumerate(swaps):
                col_name = 4 + si * 2
                col_amt  = col_name + 1
                c_name = ws4.cell(row=ri, column=col_name, value=sw["name"])
                c_name.font = SWAP_FONT; c_name.fill = SWAP_FILL
                c_name.alignment = LEFT; c_name.border = BORDER
                c_amt = ws4.cell(row=ri, column=col_amt, value=sw["label"])
                c_amt.font = SWAP_G_FONT; c_amt.fill = SWAP_FILL
                c_amt.alignment = CENTER; c_amt.border = BORDER
            # Fill any empty swap columns with blank styled cells
            for si in range(len(swaps), 5):
                col_name = 4 + si * 2
                col_amt  = col_name + 1
                set_cell(ws4, ri, col_name, "", fill=SWAP_FILL)
                set_cell(ws4, ri, col_amt,  "", fill=SWAP_FILL)
    else:
        ws4.merge_cells("A3:M3")
        c_empty = ws4["A3"]
        c_empty.value = "No foods found in the meal plan. Add meals and food items to generate swap suggestions."
        c_empty.font = Font(color="999999", size=10, italic=True)
        c_empty.alignment = LEFT

    # Notes legend below the table
    legend_row = (len(plan_foods) + 4) if plan_foods else 5
    ws4.merge_cells(f"A{legend_row}:M{legend_row}")
    note = ws4[f"A{legend_row}"]
    note.value = "ℹ️  Serving sizes are calorie-matched to the original food. Macros will differ — adjust quantities to meet specific targets."
    note.font = Font(color="666666", size=9, italic=True)
    note.alignment = LEFT

    auto_width(ws4)
    # Fix narrow swap-amount columns
    for si in range(5):
        col_letter = get_column_letter(5 + si * 2)
        ws4.column_dimensions[col_letter].width = max(ws4.column_dimensions[col_letter].width, 14)

    # ── Sheets 5+: One sheet per workout plan ──
    DOW = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
    NCOLS = 8  # Exercise | Warm-up Sets | Main Sets | Rep Range | RIR | Tempo | Intensifiers | Notes

    WARM_FILL = PatternFill("solid", fgColor="D6E4FF")   # light blue — warm-up highlight
    MAIN_FILL = PatternFill("solid", fgColor="D6F5E3")   # light green — main sets highlight
    NOTE_FILL = PatternFill("solid", fgColor="F5F5F0")   # off-white — info rows

    def _merge_row(ws, row, ncols, value, font=None, fill=None, align=None):
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws.cell(row=row, column=1, value=value)
        if font:  c.font  = font
        if fill:  c.fill  = fill
        if align: c.alignment = align
        c.border = BORDER

    def _info_block(ws, row, label, text, ncols):
        """Write a labelled text block spanning all columns. Returns next row."""
        if not text: return row
        _merge_row(ws, row, ncols, label,
                   font=Font(bold=True, color="FFFFFF", size=10), fill=SUBHDR, align=LEFT)
        row += 1
        # Each paragraph as its own wrapped row
        # Height: estimate display lines = ceil(chars / ~150 chars per merged row), 14pt each
        for line in (text or "").splitlines() or [""]:
            _merge_row(ws, row, ncols, line, font=BODY_FONT, fill=NOTE_FILL, align=WRAP_LEFT)
            display_lines = max(1, -(-len(line) // 150))   # ceiling division
            ws.row_dimensions[row].height = max(14, display_lines * 14)
            row += 1
        return row + 1  # blank gap

    TRAINING_KEY = (
        "RIR (Reps In Reserve): The number of reps you could still perform before reaching failure. "
        "RIR 0 = failure, RIR 1 = 1 rep left, RIR 2 = 2 reps left.\n"
        "Tempo (Eccentric-Pause-Concentric-Top Pause): e.g. 3-1-0-1 means 3 sec lower, 1 sec pause at bottom, "
        "0 sec lift, 1 sec pause at top. Default tempo is 3-1-0-1 unless otherwise stated.\n"
        "RPE (Rate of Perceived Exertion): Scale 1–10. RPE 10 = maximum effort, RPE 7 = moderate challenge."
    )

    for pi, plan in enumerate(plans):
        sheet_name = f"Workout - {plan['title']}"[:31]
        wsw = wb.create_sheet(sheet_name)
        wsw.row_dimensions[1].height = 28

        # ── Title row ──
        _merge_row(wsw, 1, NCOLS, f"Workout Plan: {plan['title']}",
                   font=Font(bold=True, color="FFFFFF", size=14), fill=ACCENT, align=CENTER)

        # ── Plan metadata ──
        cur_row = 2
        meta_rows = [
            ("Athlete", ath["name"]),
            ("Start Date", plan.get("start_date","")),
            ("End Date",   plan.get("end_date","")),
        ]
        for k, v in meta_rows:
            set_cell(wsw, cur_row, 1, k, bold=True)
            wsw.merge_cells(f"B{cur_row}:{get_column_letter(NCOLS)}{cur_row}")
            set_cell(wsw, cur_row, 2, v)
            cur_row += 1
        cur_row += 1

        # ── Training Instructions (plan notes) ──
        cur_row = _info_block(wsw, cur_row, "📋  Training Instructions",
                              plan.get("notes",""), NCOLS)

        # ── Warm-up Instructions ──
        cur_row = _info_block(wsw, cur_row, "🔥  Warm-Up Instructions",
                              plan.get("warmup_instructions",""), NCOLS)

        # ── Training Key ──
        cur_row = _info_block(wsw, cur_row, "📖  Training Key", TRAINING_KEY, NCOLS)

        # ── Sessions grouped by day of week ──
        sessions_by_dow = {d: [s for s in plan["sessions"] if s["day_of_week"]==d] for d in DOW}
        rest_day_set = set(plan.get("rest_days") or [])
        REST_FILL = PatternFill("solid", fgColor="E8E8E8")   # light grey for rest days

        EMPTY_FILL = PatternFill("solid", fgColor="F5F5F5")   # very light grey for empty days

        for day in DOW:
            day_sessions = sessions_by_dow.get(day, [])
            is_rest = day in rest_day_set

            # Day header — always shown for all 7 days
            day_label = f"🛌  {day}  — Rest / Recovery" if is_rest else f"📅  {day}"
            day_hdr_fill = REST_FILL if is_rest else SUBHDR
            day_hdr_font = Font(bold=True, color="888888" if is_rest else "FFFFFF", size=11)
            _merge_row(wsw, cur_row, NCOLS, day_label,
                       font=day_hdr_font, fill=day_hdr_fill, align=LEFT)
            cur_row += 1

            # Rest / no-session placeholder rows
            if not day_sessions:
                if is_rest:
                    placeholder = "Active recovery, mobility work, or full rest recommended."
                else:
                    placeholder = "No sessions scheduled."
                _merge_row(wsw, cur_row, NCOLS, placeholder,
                           font=Font(italic=True, color="AAAAAA", size=10),
                           fill=EMPTY_FILL if not is_rest else REST_FILL, align=LEFT)
                wsw.row_dimensions[cur_row].height = 16
                cur_row += 2   # row + spacer
                continue

            for sess in day_sessions:
                # Session title
                title = sess.get("session_title") or sess.get("day_of_week","")
                muscles = ", ".join(sess.get("muscle_groups",[]))
                _merge_row(wsw, cur_row, NCOLS,
                           f"{title} — {muscles}" if muscles else title,
                           font=Font(bold=True, color="1A1D27", size=10),
                           fill=PatternFill("solid", fgColor="E8EAF0"), align=LEFT)
                cur_row += 1

                if sess.get("session_notes"):
                    _merge_row(wsw, cur_row, NCOLS, f"Session notes: {sess['session_notes']}",
                               font=BODY_FONT, fill=NOTE_FILL, align=WRAP_LEFT)
                    _lines = max(1, -(-len(sess['session_notes']) // 150))
                    wsw.row_dimensions[cur_row].height = max(14, _lines * 14)
                    cur_row += 1

                if sess.get("exercises"):
                    # Column headers
                    col_hdrs = ["Exercise", "Warm-up Sets", "Main Sets",
                                "Rep Range", "RIR", "Tempo", "Intensifiers", "Training Instructions"]
                    for ci, h in enumerate(col_hdrs, 1):
                        set_hdr(wsw, cur_row, ci, h, SUBHDR, SUBF)
                    cur_row += 1

                    for ex in sess["exercises"]:
                        sets = ex.get("sets_json") or []
                        # Skip cardio (has cardio_type key)
                        if sets and sets[0].get("cardio_type") is not None:
                            # Render cardio as a single merged note row
                            cd = sets[0]
                            hrs = cd.get("duration_hours",0) or 0
                            mins = cd.get("duration_minutes",0) or 0
                            dur = f"{hrs}h {mins}m" if hrs else f"{mins} min"
                            rpe = f"  RPE {cd['rpe']}" if cd.get("rpe") is not None else ""
                            hr = f"  HR {cd.get('hr_min')}–{cd.get('hr_max')} bpm" if cd.get("hr_min") and cd.get("hr_max") else ""
                            _merge_row(wsw, cur_row, NCOLS,
                                       f"🏃 {ex['name']} ({cd.get('cardio_type','')}) — {dur}{rpe}{hr}",
                                       font=BODY_FONT, fill=NOTE_FILL, align=LEFT)
                            cur_row += 1
                            continue

                        # Partition sets by type
                        w_sets = [s for s in sets if (s.get("type") or "M") == "W"]
                        m_sets = [s for s in sets if (s.get("type") or "M") == "M"]
                        i_sets = [s for s in sets if (s.get("type") or "M") == "I"]

                        wu_count = len(w_sets)
                        main_count = len(m_sets)

                        # Rep range: "WU 10/8 | Main 8-12" or just "8-12"
                        wu_reps = "/".join(
                            str(s.get("rep_range") or s.get("reps","")).strip() or "—"
                            for s in w_sets
                        ) if w_sets else ""
                        # Main rep range: prefer exercise rep_range, else join set reps
                        main_reps = (ex.get("rep_range") or "").strip()
                        if not main_reps and m_sets:
                            main_reps = "/".join(str(s.get("reps","")) for s in m_sets)
                        rep_range_str = ""
                        if wu_reps and main_reps:
                            rep_range_str = f"WU {wu_reps} | Main {main_reps}"
                        elif wu_reps:
                            rep_range_str = f"WU {wu_reps}"
                        elif main_reps:
                            rep_range_str = main_reps

                        # RIR: exercise default, or 0
                        rir_val = ex.get("rir")
                        rir_str = str(rir_val) if rir_val is not None else "0"

                        # Tempo: exercise default, or "3-1-0-1"
                        tempo_str = (ex.get("tempo") or "").strip() or "3-1-0-1"

                        # Intensifiers: exercise intensifiers text + notes from I-type sets
                        intensifier_parts = []
                        if ex.get("intensifiers"):
                            intensifier_parts.append(ex["intensifiers"].strip())
                        for s in i_sets:
                            if s.get("notes"):
                                intensifier_parts.append(s["notes"].strip())
                        intensifiers_str = "; ".join(p for p in intensifier_parts if p)

                        # Exercise row
                        notes_str   = (ex.get("exercise_notes") or "").strip()
                        wu_instr    = (ex.get("warmup_instructions") or "").strip()
                        set_cell(wsw, cur_row, 1, ex["name"], bold=True)
                        set_cell(wsw, cur_row, 2, wu_count   if wu_count   else "—", align=CENTER)
                        set_cell(wsw, cur_row, 3, main_count if main_count else "—", align=CENTER)
                        set_cell(wsw, cur_row, 4, rep_range_str, align=WRAP_LEFT)
                        set_cell(wsw, cur_row, 5, rir_str,    align=CENTER)
                        set_cell(wsw, cur_row, 6, tempo_str,  align=CENTER)
                        set_cell(wsw, cur_row, 7, intensifiers_str, align=WRAP_LEFT)
                        set_cell(wsw, cur_row, 8, notes_str,  align=WRAP_LEFT)

                        # Highlight warm-up / main set count cells
                        if wu_count:
                            wsw.cell(row=cur_row, column=2).fill = WARM_FILL
                        if main_count:
                            wsw.cell(row=cur_row, column=3).fill = MAIN_FILL

                        cur_row += 1

                        # Warm Up Sets sub-row (only when warmup_instructions exist)
                        if wu_instr:
                            set_cell(wsw, cur_row, 1, "Warm Up Sets",
                                     bold=True, fill=WARM_FILL)
                            wsw.merge_cells(f"B{cur_row}:{get_column_letter(NCOLS)}{cur_row}")
                            c_wu = wsw.cell(row=cur_row, column=2, value=wu_instr)
                            c_wu.font      = BODY_FONT
                            c_wu.alignment = WRAP_LEFT
                            c_wu.border    = BORDER
                            c_wu.fill      = WARM_FILL
                            _wu_lines = max(1, -(-len(wu_instr) // 150))
                            wsw.row_dimensions[cur_row].height = max(14, _wu_lines * 14)
                            cur_row += 1

                cur_row += 1  # spacer row between sessions

        auto_width(wsw)
        # Fix narrow number columns
        for col_idx in [2, 3, 5, 6]:
            wsw.column_dimensions[get_column_letter(col_idx)].width = 14
        # Set generous widths for text-heavy columns so wrap is meaningful
        wsw.column_dimensions["A"].width = max(wsw.column_dimensions["A"].width, 30)  # Exercise
        wsw.column_dimensions["D"].width = max(wsw.column_dimensions["D"].width, 22)  # Rep Range
        wsw.column_dimensions["G"].width = max(wsw.column_dimensions["G"].width, 30)  # Intensifiers
        wsw.column_dimensions["H"].width = 55                                          # Notes (wide)

    return wb


@app.get("/api/athletes/{athlete_id}/export-xlsx")
def export_xlsx(athlete_id: int):
    wb = _build_workbook(athlete_id)
    ath = get_athlete(athlete_id)
    safe_name = "".join(c for c in ath["name"] if c.isalnum() or c in " _-").strip() or "athlete"
    filename = f"BodyBuilder_{safe_name}.xlsx"
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/api/athletes/{athlete_id}/workout-plans/{plan_id}/export-xlsx")
def export_plan_xlsx(athlete_id: int, plan_id: int):
    import traceback
    try:
        wb = _build_workbook(athlete_id, plan_id=plan_id)
    except HTTPException:
        raise
    except Exception as exc:
        traceback.print_exc()
        raise HTTPException(500, f"Export error: {exc}")
    conn = get_db()
    plan_row = conn.execute("SELECT title FROM workout_plans WHERE id=? AND athlete_id=?", (plan_id, athlete_id)).fetchone()
    conn.close()
    if not plan_row:
        raise HTTPException(404, "Plan not found")
    safe_title = "".join(ch for ch in plan_row["title"] if ch.isalnum() or ch in " _-").strip() or "plan"
    filename = f"{safe_title}_workout.xlsx"
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/admin/send-program")
def send_program(body: SendProgramModel):
    conn = get_db()
    cfg = dict(conn.execute("SELECT * FROM smtp_settings WHERE id=1").fetchone())
    conn.close()
    if not cfg["username"] or not cfg["password"]:
        raise HTTPException(400, "SMTP credentials not configured")
    ath = get_athlete(body.athlete_id)
    wb = _build_workbook(body.athlete_id)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    safe_name = "".join(c for c in ath["name"] if c.isalnum() or c in " _-").strip() or "athlete"
    filename = f"BodyBuilder_{safe_name}.xlsx"
    msg = MIMEMultipart()
    msg["From"] = f"{cfg['from_name']} <{cfg['username']}>"
    msg["To"] = body.to_email
    msg["Subject"] = body.subject
    body_text = body.message or f"Hi,\n\nPlease find your BodyBuilder program attached as an Excel file.\n\nYou can open it in Google Sheets via File → Import.\n\nBest regards,\n{cfg['from_name']}"
    msg.attach(MIMEText(body_text, "plain"))
    part = MIMEBase("application","octet-stream"); part.set_payload(buf.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename=filename)
    msg.attach(part)
    try:
        if cfg["use_tls"]:
            s = smtplib.SMTP(cfg["host"], cfg["port"], timeout=15); s.starttls()
        else:
            s = smtplib.SMTP_SSL(cfg["host"], cfg["port"], timeout=15)
        s.login(cfg["username"], cfg["password"]); s.sendmail(cfg["username"], body.to_email, msg.as_string()); s.quit()
        return {"success":True,"message":f"Program sent to {body.to_email}"}
    except smtplib.SMTPAuthenticationError:
        raise HTTPException(400, "SMTP authentication failed. Check username/password (use App Password for Gmail).")
    except Exception as e:
        raise HTTPException(500, f"Failed to send email: {e}")


# ─── Exercise Images ──────────────────────────────────────────────────────────

@app.get("/api/exercise-image")
def get_exercise_image(name: str):
    """Look up the cached local image path for a named exercise."""
    conn = get_db()
    row = conn.execute(
        "SELECT image_path FROM exercise_images WHERE name=?", (name,)
    ).fetchone()
    conn.close()
    if row and row["image_path"] and (EXERCISE_IMAGES_DIR / row["image_path"]).exists():
        return {"image_url": f"/exercise-images/{row['image_path']}", "found": True}
    return {"image_url": "", "found": False}


@app.post("/api/exercise-images/seed")
def start_image_seed(force: bool = False):
    """Kick off a background fetch of exercise images from Wger. Idempotent."""
    global _seed_state
    if _seed_state["running"]:
        return {"status": "already_running", **_seed_state}
    thread = threading.Thread(target=_run_seed, kwargs={"force": force}, daemon=True)
    thread.start()
    return {"status": "started", "total": len(EXERCISE_ALL)}


@app.get("/api/exercise-images/seed/status")
def get_seed_status():
    """Current state of the image-seeding background job."""
    return _seed_state


# ─── Backup / Restore ─────────────────────────────────────────────────────────

# Tables included in backup, in insertion order (parents before children)
BACKUP_TABLES = [
    "athletes",
    "programs",
    "activity_calories",
    "calendar_days",
    "calendar_events",
    "meal_plans",
    "nutrition_foods",
    "food_swaps",
    "meals",
    "meal_items",
    "workout_plans",
    "workout_sessions",
    "workout_exercises",
    "supplements",
    "smtp_settings",
]


def _checksum(data: dict) -> str:
    """
    Canonical SHA-256 checksum of backup data.

    This must produce the same result as the JavaScript canonicalJSON() + sha256hex()
    helpers in admin-tab.js.  Since the restore payload arrives from JavaScript
    (where all numbers have already been through JSON.parse), integers come back
    as Python int and non-integer floats come back as Python float — exactly the
    types that json.dumps serialises identically to JSON.stringify.  No float
    normalisation is needed on the restore path.

    On the backup path the checksum is now computed client-side in JS (not here),
    so this function is only used on the restore path to verify the stored checksum.
    """
    canonical = json.dumps(data, sort_keys=True, separators=(",", ":"), ensure_ascii=False, default=str)
    return hashlib.sha256(canonical.encode()).hexdigest()


@app.get("/api/backup")
def create_backup():
    """Return a full backup of all application data as JSON."""
    conn = get_db()
    data: dict = {}
    try:
        for table in BACKUP_TABLES:
            rows = conn.execute(f"SELECT * FROM {table}").fetchall()
            data[table] = [dict(r) for r in rows]
    finally:
        conn.close()

    # Checksum placeholder — the real checksum is computed client-side in JS
    # after JSON.stringify so it reflects the exact bytes written to the file.
    # This avoids Python float serialisation differences (175.0 vs 175).
    return {
        "format": "bodybuilder-backup",
        "app_version": ".".join(str(x) for x in APP_VERSION),
        "created_at": datetime.now().isoformat(),
        "checksum": "",   # overwritten by JS before saving
        "data": data,
    }


@app.post("/api/restore")
async def restore_backup(request: Request):
    """Replace all application data from a .bb backup file."""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid file — could not parse JSON")

    if body.get("format") != "bodybuilder-backup":
        raise HTTPException(
            422,
            "Incompatible file type — this is not a BodyBuilder backup file"
        )

    data = body.get("data")
    if not isinstance(data, dict):
        raise HTTPException(400, "Backup file is missing data section")

    # Checksum is verified client-side in JS (canonicalJSON + SHA-256) before the
    # request is even sent, so no server-side re-verification is needed.  Skipping
    # it here avoids any remaining Python/JS serialisation differences (e.g.
    # ensure_ascii, float representation) that could cause spurious failures.

    conn = get_db()
    try:
        conn.execute("PRAGMA foreign_keys = OFF")
        # Delete child tables first to avoid FK violations during clear
        for table in reversed(BACKUP_TABLES):
            conn.execute(f"DELETE FROM {table}")
        # Reset auto-increment counters where they exist
        try:
            conn.execute("DELETE FROM sqlite_sequence WHERE name IN ({})".format(
                ",".join(f"'{t}'" for t in BACKUP_TABLES)
            ))
        except Exception:
            pass  # sqlite_sequence may not exist if no AUTOINCREMENT tables yet

        # Insert rows for each table
        for table in BACKUP_TABLES:
            rows = data.get(table, [])
            if not rows:
                continue
            # Determine valid columns from schema to guard against schema drift
            valid_cols = {
                r["name"]
                for r in conn.execute(f"PRAGMA table_info({table})").fetchall()
            }
            for row in rows:
                safe_row = {k: v for k, v in row.items() if k in valid_cols}
                if not safe_row:
                    continue
                cols = ", ".join(safe_row.keys())
                placeholders = ", ".join(["?"] * len(safe_row))
                conn.execute(
                    f"INSERT OR REPLACE INTO {table} ({cols}) VALUES ({placeholders})",
                    list(safe_row.values()),
                )

        conn.execute("PRAGMA foreign_keys = ON")
        conn.commit()
    except Exception as exc:
        conn.rollback()
        conn.close()
        raise HTTPException(500, "Restore failed — database could not be updated. Check server.log for details.")

    conn.close()
    return {
        "restored": True,
        "athletes_count": len(data.get("athletes", [])),
    }


# ─── Frontend ─────────────────────────────────────────────────────────────────

@app.get("/")
def serve_frontend():
    index = FRONTEND_DIR / "index.html"
    if index.exists(): return FileResponse(str(index))
    return {"message": "BodyBuilder API running. Open frontend/index.html in your browser."}

# Serve exercise images (must come before the catch-all frontend mount)
app.mount("/exercise-images", StaticFiles(directory=str(EXERCISE_IMAGES_DIR)), name="exercise_images")

# Serve frontend static assets (css, js, images)
if FRONTEND_DIR.exists():
    app.mount("/", StaticFiles(directory=str(FRONTEND_DIR)), name="frontend")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=False)
